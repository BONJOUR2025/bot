"""API endpoints for payroll calculation."""
from __future__ import annotations

from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.services.payroll_service import PayrollService, get_payroll_service, make_month_key
from app.data.sales_plans_repository import SalesPlansRepository, get_sales_plans_repository
from app.data.payroll_settlement_repository import get_payroll_settlement_repository
from app.data.payroll_audit_repository import get_payroll_audit_repository
from app.data.payroll_comments_repository import get_payroll_comments_repository
from app.services.access_control_service import AccessControlService, ResolvedUser
from app.services.employee_service import EmployeeService
from .dependencies import get_current_user


class SalesPlanInput(BaseModel):
    employee_code: str
    employee_name: str
    month_key: str | None = None
    repair_plan: float | None = None
    cosmetics_plan: float | None = None
    shoes_plan: float | None = None
    ignore_kpi: bool | None = None
    force_max: list = []
    force_min: list = []


class SalesPlanOutput(BaseModel):
    employee_code: str
    employee_name: str
    month_key: str | None = None
    repair_plan: float
    cosmetics_plan: float
    shoes_plan: float
    ignore_kpi: bool
    force_max: list = []
    force_min: list = []


class PayrollRowOutput(BaseModel):
    employee_code: str
    employee_name: str
    base_salary: float
    repair_sales: float
    cosmetics_sales: float
    shoes_sales: float
    repair_plan: float
    cosmetics_plan: float
    shoes_plan: float
    repair_fulfillment: float
    cosmetics_fulfillment: float
    shoes_fulfillment: float
    repair_rate: float
    cosmetics_rate: float
    shoes_rate: float
    repair_commission: float
    cosmetics_commission: float
    shoes_commission: float
    bonuses: float
    excel_bonus: float
    penalties: float
    advances: float
    advances_this_month: float = 0.0
    ignore_kpi: bool
    force_max: list = []
    force_min: list = []
    shoes_orders: list[str]
    total_commission: float
    total_gross: float
    total_deductions: float
    total_net: float
    settlement_paid: bool = False
    shifts_by_point: dict = {}


class SettlementInput(BaseModel):
    paid: bool


class AuditEntryInput(BaseModel):
    action: str
    employee_code: str
    employee_name: str
    month_key: str
    details: dict[str, Any] = {}


class CommentInput(BaseModel):
    comment: str


class SaleTransferInput(BaseModel):
    month: str
    year: int
    doc_num: str
    from_category: str          # repair / cosmetics / shoes
    to_category: str            # repair / cosmetics / shoes
    amount: float = 0.0
    from_code: str
    to_code: str
    from_name: str = ""
    to_name: str = ""
    order_date: str = ""
    shoes_orders: list = []


def create_payroll_router(
    payroll_service: PayrollService | None = None,
    plans_repo: SalesPlansRepository | None = None,
    access_service: AccessControlService | None = None,
    employee_service: EmployeeService | None = None,
) -> APIRouter:
    router = APIRouter(prefix="/payroll", tags=["Payroll"])

    _payroll = payroll_service or get_payroll_service()
    _plans = plans_repo or get_sales_plans_repository()
    _settlements = get_payroll_settlement_repository()
    _audit = get_payroll_audit_repository()
    _comments = get_payroll_comments_repository()
    _employees = employee_service or EmployeeService()

    def _check(current: ResolvedUser) -> None:
        if access_service and not access_service.user_has_permission(current, "payroll"):
            raise HTTPException(status_code=403, detail="forbidden")

    # ── Employee self-service ──────────────────────────────────────
    @router.get("/my", response_model=PayrollRowOutput | None)
    async def get_my_payroll(
        month: str = Query(...),
        year: Optional[int] = Query(None),
        current: ResolvedUser = Depends(get_current_user),
    ):
        """Returns payroll data for the currently authenticated employee."""
        if not current.employee_id:
            raise HTTPException(status_code=403, detail="Not an employee account")
        emp = _employees.get_employee(current.employee_id)
        full_name = emp.full_name if emp else None
        code = _payroll.get_code_for_employee(
            employee_id=current.employee_id, full_name=full_name
        )
        if not code:
            return None
        row = await _payroll.get_employee_details(code, month, year)
        return PayrollRowOutput(**row.to_dict()) if row else None

    # ── Months ────────────────────────────────────────────────────
    @router.get("/months", response_model=list[str])
    async def list_months(current: ResolvedUser = Depends(get_current_user)):
        _check(current)
        return _payroll.list_months()

    # ── Calculate ─────────────────────────────────────────────────
    @router.get("/calculate")
    async def calculate_payroll(
        month: str = Query(...),
        year: Optional[int] = Query(None),
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        rows, unknown_codes = await _payroll.calculate_payroll(month, year)
        return {
            "rows": [PayrollRowOutput(**row.to_dict()) for row in rows],
            "unknown_codes": unknown_codes,
        }

    @router.get("/calculate/{employee_code}", response_model=PayrollRowOutput | None)
    async def get_employee_payroll(
        employee_code: str,
        month: str = Query(...),
        year: Optional[int] = Query(None),
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        row = await _payroll.get_employee_details(employee_code, month, year)
        return PayrollRowOutput(**row.to_dict()) if row else None

    # ── Advances history ──────────────────────────────────────────
    @router.get("/advances-history")
    async def advances_history(
        month: str = Query(...),
        year: int = Query(...),
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        return _payroll.get_advances_history(month, year)

    # ── Settlements ───────────────────────────────────────────────
    @router.get("/settlements")
    async def get_settlements(
        month: str = Query(...),
        year: int = Query(...),
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        return _settlements.get_settlements_map(make_month_key(month, year))

    @router.put("/settlements/{employee_code}")
    async def set_settlement(
        employee_code: str,
        data: SettlementInput,
        month: str = Query(...),
        year: int = Query(...),
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        return _settlements.set_settlement(make_month_key(month, year), employee_code, data.paid)

    # ── Plans ─────────────────────────────────────────────────────
    @router.get("/plans", response_model=list[SalesPlanOutput])
    async def list_plans(
        month: Optional[str] = Query(None),
        year: Optional[int] = Query(None),
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        if month and year:
            mk = make_month_key(month, year)
            month_plans = _plans.list_plans(month_key=mk)
            month_codes = {p.employee_code for p in month_plans}
            # Include global plans not overridden by month-specific ones
            global_plans = [
                p for p in _plans.list_plans(month_key=None)
                if p.month_key is None and p.employee_code not in month_codes
            ]
            all_plans = month_plans + global_plans
        else:
            all_plans = _plans.list_plans()

        return [
            SalesPlanOutput(
                employee_code=p.employee_code,
                employee_name=p.employee_name,
                month_key=p.month_key,
                repair_plan=p.repair_plan,
                cosmetics_plan=p.cosmetics_plan,
                shoes_plan=p.shoes_plan,
                ignore_kpi=p.ignore_kpi,
                force_max=p.force_max,
                force_min=p.force_min,
            )
            for p in all_plans
        ]

    @router.get("/plans/{employee_code}", response_model=SalesPlanOutput | None)
    async def get_plan(
        employee_code: str,
        month: Optional[str] = Query(None),
        year: Optional[int] = Query(None),
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        mk = make_month_key(month, year) if month and year else None
        plan = _plans.get_plan(employee_code, month_key=mk)
        if not plan:
            return None
        return SalesPlanOutput(
            employee_code=plan.employee_code,
            employee_name=plan.employee_name,
            month_key=plan.month_key,
            repair_plan=plan.repair_plan,
            cosmetics_plan=plan.cosmetics_plan,
            shoes_plan=plan.shoes_plan,
            ignore_kpi=plan.ignore_kpi,
            force_max=plan.force_max,
            force_min=plan.force_min,
        )

    @router.put("/plans", response_model=SalesPlanOutput)
    async def set_plan(
        data: SalesPlanInput,
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        plan = _plans.set_plan(
            employee_code=data.employee_code,
            employee_name=data.employee_name,
            month_key=data.month_key,
            repair_plan=data.repair_plan,
            cosmetics_plan=data.cosmetics_plan,
            shoes_plan=data.shoes_plan,
            ignore_kpi=data.ignore_kpi,
            force_max=data.force_max,
            force_min=data.force_min,
        )
        return SalesPlanOutput(
            employee_code=plan.employee_code,
            employee_name=plan.employee_name,
            month_key=plan.month_key,
            repair_plan=plan.repair_plan,
            cosmetics_plan=plan.cosmetics_plan,
            shoes_plan=plan.shoes_plan,
            ignore_kpi=plan.ignore_kpi,
            force_max=plan.force_max,
            force_min=plan.force_min,
        )

    @router.delete("/plans/{employee_code}")
    async def delete_plan(
        employee_code: str,
        month: Optional[str] = Query(None),
        year: Optional[int] = Query(None),
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        mk = make_month_key(month, year) if month and year else None
        if not _plans.delete_plan(employee_code, month_key=mk):
            raise HTTPException(status_code=404, detail="Plan not found")
        return {"status": "deleted"}

    # ── Audit log ──────────────────────────────────────────────────
    @router.get("/audit")
    async def get_audit_log(
        month: Optional[str] = Query(None),
        year: Optional[int] = Query(None),
        employee_code: Optional[str] = Query(None),
        limit: int = Query(200),
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        mk = make_month_key(month, year) if month and year else None
        return _audit.get_entries(month_key=mk, employee_code=employee_code, limit=limit)

    @router.post("/audit")
    async def add_audit_entry(
        data: AuditEntryInput,
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        actor = getattr(current, "login", None) or getattr(current, "id", "admin")
        entry = _audit.add_entry(
            actor=str(actor),
            action=data.action,
            employee_code=data.employee_code,
            employee_name=data.employee_name,
            month_key=data.month_key,
            details=data.details,
        )
        return entry

    # ── Comments ───────────────────────────────────────────────────
    @router.get("/comments")
    async def get_comments(
        month: str = Query(...),
        year: int = Query(...),
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        return _comments.get_comments_map(make_month_key(month, year))

    @router.put("/comments/{employee_code}")
    async def set_comment(
        employee_code: str,
        data: CommentInput,
        month: str = Query(...),
        year: int = Query(...),
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        mk = make_month_key(month, year)
        if data.comment.strip():
            return _comments.set_comment(mk, employee_code, data.comment.strip())
        else:
            _comments.delete_comment(mk, employee_code)
            return {"status": "deleted"}

    # ── Sale transfers (move an order's sale between employees) ─────
    @router.get("/order-lookup")
    async def order_lookup(
        doc_num: str = Query(...),
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        doc_num = doc_num.strip()
        if not doc_num:
            raise HTTPException(status_code=400, detail="empty_doc_num")
        breakdown = _payroll.firebird.get_order_breakdown(doc_num)
        if not breakdown.get("found"):
            raise HTTPException(status_code=404, detail="order_not_found")
        return breakdown

    @router.get("/sale-transfers")
    async def list_sale_transfers(
        month: str = Query(...),
        year: int = Query(...),
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        from app.services.sale_transfer_service import list_transfers
        return list_transfers(make_month_key(month, year))

    @router.post("/sale-transfers")
    async def create_sale_transfer(
        data: SaleTransferInput,
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        from app.services.sale_transfer_service import create_transfer
        author = getattr(current, "login", None) or getattr(current, "id", "admin")
        try:
            return create_transfer(
                month_key=make_month_key(data.month, data.year),
                doc_num=data.doc_num,
                from_category=data.from_category,
                to_category=data.to_category,
                amount=data.amount,
                from_code=data.from_code,
                to_code=data.to_code,
                from_name=data.from_name,
                to_name=data.to_name,
                order_date=data.order_date,
                shoes_orders=data.shoes_orders,
                author=str(author),
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))

    @router.delete("/sale-transfers/{transfer_id}")
    async def delete_sale_transfer(
        transfer_id: int,
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        from app.services.sale_transfer_service import delete_transfer
        if not delete_transfer(transfer_id):
            raise HTTPException(status_code=404, detail="not_found")
        return {"status": "deleted"}

    # ── Excel export ───────────────────────────────────────────────
    @router.get("/export/excel")
    async def export_excel(
        month: str = Query(...),
        year: Optional[int] = Query(None),
        current: ResolvedUser = Depends(get_current_user),
    ):
        _check(current)
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        rows, _ = await _payroll.calculate_payroll(month, year)
        actual_year = year or __import__("datetime").date.today().year
        month_key = make_month_key(month, actual_year)
        settlements = _settlements.get_settlements_map(month_key)
        comments_map = _comments.get_comments_map(month_key)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{month} {actual_year}"

        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        alt_fill    = PatternFill("solid", fgColor="F0F4F8")
        paid_fill   = PatternFill("solid", fgColor="D1FAE5")
        warn_fill   = PatternFill("solid", fgColor="FEF3C7")
        center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
        right       = Alignment(horizontal="right",  vertical="center")
        left        = Alignment(horizontal="left",   vertical="center")

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = [
            "ФИО", "Оклад", "Комиссия", "Ремонт %", "Косм. %",
            "Обувь %", "Премии", "Авансы", "Штрафы", "К выплате",
            "Выдано", "Комментарий",
        ]
        col_widths = [22, 14, 14, 10, 10, 10, 12, 12, 12, 16, 10, 30]

        for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.row_dimensions[1].height = 30

        def pct(v):
            return f"{(v * 100):.0f}%" if v is not None else "—"

        def rub(v):
            return round(v) if v else 0

        for ri, row in enumerate(rows, start=2):
            fill = paid_fill if settlements.get(row.employee_code) else (alt_fill if ri % 2 == 0 else None)

            deduction_ratio = (row.advances + row.penalties) / row.total_gross if row.total_gross else 0
            low_repair = row.repair_fulfillment is not None and row.repair_fulfillment < 0.5 and row.repair_plan > 0
            low_cosm   = row.cosmetics_fulfillment is not None and row.cosmetics_fulfillment < 0.5 and row.cosmetics_plan > 0
            has_anomaly = deduction_ratio > 0.2 or low_repair or low_cosm or row.force_max or row.force_min
            if has_anomaly and not settlements.get(row.employee_code):
                fill = warn_fill

            values = [
                row.employee_name,
                rub(row.base_salary),
                rub(row.total_commission) if not row.ignore_kpi else 0,
                pct(row.repair_fulfillment),
                pct(row.cosmetics_fulfillment),
                pct(row.shoes_fulfillment) if hasattr(row, "shoes_fulfillment") else "—",
                rub(row.bonuses + row.excel_bonus),
                rub(row.advances),
                rub(row.penalties),
                rub(row.total_net),
                "Да" if settlements.get(row.employee_code) else "Нет",
                comments_map.get(row.employee_code, ""),
            ]
            alignments = [left, right, right, center, center, center, right, right, right, right, center, left]

            for ci, (val, aln) in enumerate(zip(values, alignments), start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = aln
                cell.border = border
                if fill:
                    cell.fill = fill

        # Totals row
        tr = len(rows) + 2
        ws.cell(row=tr, column=1, value="ИТОГО").font = Font(bold=True)
        for ci in range(1, len(headers) + 1):
            ws.cell(row=tr, column=ci).border = border
        ws.cell(row=tr, column=2, value=sum(rub(r.base_salary) for r in rows)).font = Font(bold=True)
        ws.cell(row=tr, column=3, value=sum(rub(r.total_commission) for r in rows if not r.ignore_kpi)).font = Font(bold=True)
        ws.cell(row=tr, column=7, value=sum(rub(r.bonuses + r.excel_bonus) for r in rows)).font = Font(bold=True)
        ws.cell(row=tr, column=8, value=sum(rub(r.advances) for r in rows)).font = Font(bold=True)
        ws.cell(row=tr, column=9, value=sum(rub(r.penalties) for r in rows)).font = Font(bold=True)
        ws.cell(row=tr, column=10, value=sum(rub(r.total_net) for r in rows)).font = Font(bold=True)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"payroll_{month}_{actual_year}.xlsx"
        headers_resp = {
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_resp)

    return router
