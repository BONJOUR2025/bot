"""Payroll calculation service - combines Excel, Firebird, advances and bonuses."""
from __future__ import annotations

import asyncio
import calendar
import json
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from app.settings import settings
from app.data.sales_plans_repository import SalesPlansRepository, get_sales_plans_repository
from app.data.payroll_settlement_repository import PayrollSettlementRepository, get_payroll_settlement_repository
from app.data.location_repository import LocationRepository, get_location_repository
from app.data.salon_repository import SalonRepository, get_salon_repository
from app.services.firebird_service import FirebirdService, get_firebird_service
from app.config import EXCEL_FILE

logger = logging.getLogger(__name__)

MONTH_NAMES = {
    "ЯНВАРЬ": 1, "ФЕВРАЛЬ": 2, "МАРТ": 3, "АПРЕЛЬ": 4,
    "МАЙ": 5, "ИЮНЬ": 6, "ИЮЛЬ": 7, "АВГУСТ": 8,
    "СЕНТЯБРЬ": 9, "ОКТЯБРЬ": 10, "НОЯБРЬ": 11, "ДЕКАБРЬ": 12,
}

REPAIR_RATE_HIGH = 0.02
REPAIR_RATE_LOW = 0.01
COSMETICS_RATE_HIGH = 0.08
COSMETICS_RATE_LOW = 0.05
SHOES_RATE_HIGH = 0.05
SHOES_RATE_LOW = 0.03
PLAN_THRESHOLD = 0.80

CODE_RE = re.compile(r"(\d{4})$")

# Firebird order numbers look like "12345-6" or "12345-12" — the digits
# after the dash identify the salon the sale happened at. Separate code
# space from the employee's 4-digit CODE_RE and from Salon.code (the
# letter-based schedule/shift point code).
ORDER_SALON_CODE_RE = re.compile(r"-(\d{1,2})$")


def _extract_code(name: str | None) -> str | None:
    m = CODE_RE.search((name or "").strip())
    return m.group(1) if m else None


def _order_salon_code(doc_num: str | None) -> str | None:
    m = ORDER_SALON_CODE_RE.search((doc_num or "").strip())
    return m.group(1) if m else None


def _parse_dt(v) -> datetime:
    if isinstance(v, datetime):
        return v
    if v is None:
        return datetime(1970, 1, 1)
    if isinstance(v, (int, float)):
        ts = float(v)
        if ts > 10_000_000_000:
            ts /= 1000.0
        return datetime.fromtimestamp(ts)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return datetime(1970, 1, 1)
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"
        try:
            return datetime.fromisoformat(s)
        except ValueError:
            try:
                return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return datetime(1970, 1, 1)
    return datetime(1970, 1, 1)


def _dt_field(r: dict) -> Any:
    return r.get("date") or r.get("timestamp") or r.get("created_at") or r.get("createdAt")


def make_month_key(month: str, year: int) -> str:
    """Build a canonical month key, e.g. 'ЯНВАРЬ_2025'."""
    return f"{month.strip().upper()}_{year}"


@dataclass
class PayrollRow:
    employee_code: str
    employee_name: str
    base_salary: float

    repair_sales: float
    cosmetics_sales: float
    shoes_sales: float

    repair_plan: float
    cosmetics_plan: float
    shoes_plan: float

    repair_fulfillment: float
    cosmetics_fulfillment: float
    shoes_fulfillment: float

    repair_rate: float
    cosmetics_rate: float
    shoes_rate: float

    repair_commission: float
    cosmetics_commission: float
    shoes_commission: float

    bonuses: float
    excel_bonus: float
    penalties: float

    advances: float              # advances since last salary (actual deduction)
    advances_this_month: float   # advances taken in this calendar month (display only)
    ignore_kpi: bool
    force_max: list
    force_min: list
    shoes_orders: list

    total_commission: float
    total_gross: float
    total_deductions: float
    total_net: float

    main_rate: float = 0.0
    main_shifts: float = 0.0
    extra_rate: float = 0.0
    extra_shifts: float = 0.0
    workshop_commission: float = 0.0
    settlement_paid: bool = False
    shifts_by_point: dict = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "employee_code": self.employee_code,
            "employee_name": self.employee_name,
            "base_salary": self.base_salary,
            "repair_sales": self.repair_sales,
            "cosmetics_sales": self.cosmetics_sales,
            "shoes_sales": self.shoes_sales,
            "repair_plan": self.repair_plan,
            "cosmetics_plan": self.cosmetics_plan,
            "shoes_plan": self.shoes_plan,
            "repair_fulfillment": self.repair_fulfillment,
            "cosmetics_fulfillment": self.cosmetics_fulfillment,
            "shoes_fulfillment": self.shoes_fulfillment,
            "repair_rate": self.repair_rate,
            "cosmetics_rate": self.cosmetics_rate,
            "shoes_rate": self.shoes_rate,
            "repair_commission": self.repair_commission,
            "cosmetics_commission": self.cosmetics_commission,
            "shoes_commission": self.shoes_commission,
            "bonuses": self.bonuses,
            "excel_bonus": self.excel_bonus,
            "penalties": self.penalties,
            "advances": self.advances,
            "advances_this_month": self.advances_this_month,
            "ignore_kpi": self.ignore_kpi,
            "force_max": self.force_max,
            "force_min": self.force_min,
            "shoes_orders": self.shoes_orders,
            "total_commission": self.total_commission,
            "total_gross": self.total_gross,
            "total_deductions": self.total_deductions,
            "total_net": self.total_net,
            "settlement_paid": self.settlement_paid,
            "shifts_by_point": self.shifts_by_point,
        }


def _calc_base_salary(shifts_total: int, main_rate: float, extra_rate: float) -> float:
    """Salary = first 15 shifts × main_rate + remaining shifts × extra_rate."""
    main_shifts  = min(shifts_total, 15)
    extra_shifts = max(0, shifts_total - 15)
    return main_shifts * main_rate + extra_shifts * extra_rate


def _parse_schedule_from_excel(month: str, year: int) -> dict[str, dict[str, int]]:
    """Read schedule sheet from EXCEL_FILE.

    Returns {employee_name: {point_code: shift_count}}.
    """
    from app.core.constants import MONTHS_RU
    month_upper = month.strip().upper()
    month_num = MONTH_NAMES.get(month_upper)
    if not month_num:
        return {}

    excel_path = Path(EXCEL_FILE)
    if not excel_path.exists():
        return {}

    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception as e:
        logger.warning(f"Cannot open schedule Excel: {e}")
        return {}

    month_name = MONTHS_RU[month_num - 1]
    sheet = None
    for candidate in (month_name, month_name.upper(), month_upper):
        if candidate in wb.sheetnames:
            sheet = wb[candidate]
            break
    if sheet is None:
        for title in wb.sheetnames:
            if title.upper().startswith(month_upper[:3]):
                sheet = wb[title]
                break
    if sheet is None:
        return {}

    # Find columns that contain day numbers (rows 1 or 2)
    days_in_month = calendar.monthrange(year, month_num)[1]
    day_cols: dict[int, int] = {}
    for col in range(1, sheet.max_column + 1):
        for row in (1, 2):
            raw = str(sheet.cell(row=row, column=col).value or "").strip()
            try:
                d = int(raw)
                if 1 <= d <= days_in_month and d not in day_cols:
                    day_cols[d] = col
            except ValueError:
                pass

    if not day_cols:
        return {}

    result: dict[str, dict[str, int]] = {}
    for row in range(3, sheet.max_row + 1):
        raw_name = sheet.cell(row=row, column=1).value
        emp_name = str(raw_name).strip() if raw_name else ""
        if not emp_name or emp_name.lower() in ("имя", "name", "nan"):
            continue
        shifts: dict[str, int] = {}
        for d, col in day_cols.items():
            val = str(sheet.cell(row=row, column=col).value or "").strip()
            if val:
                shifts[val] = shifts.get(val, 0) + 1
        if shifts:
            result[emp_name] = shifts

    return result


class PayrollService:
    def __init__(
        self,
        excel_path: str | Path | None = None,
        firebird_service: FirebirdService | None = None,
        plans_repo: SalesPlansRepository | None = None,
        settlement_repo: PayrollSettlementRepository | None = None,
        location_repo: LocationRepository | None = None,
        salon_repo: SalonRepository | None = None,
        advance_requests_file: str | None = None,
        bonuses_penalties_file: str | None = None,
        users_file: str | None = None,
    ) -> None:
        self.excel_path = Path(excel_path or settings.payroll_excel_file)
        self.firebird = firebird_service or get_firebird_service()
        self.plans_repo = plans_repo or get_sales_plans_repository()
        self.settlement_repo = settlement_repo or get_payroll_settlement_repository()
        self.location_repo = location_repo or get_location_repository()
        self.salon_repo = salon_repo or get_salon_repository()
        self.advance_requests_file = Path(advance_requests_file or settings.advance_requests_file)
        self.bonuses_penalties_file = Path(bonuses_penalties_file or settings.bonuses_penalties_file)
        self.users_file = Path(users_file or settings.users_file)

        self._user_id_to_code: dict[str, str] = {}
        self._full_name_to_code: dict[str, str] = {}
        self._users_mtime: float | None = None
        self._load_user_mappings()

    def _load_user_mappings(self) -> None:
        if not self.users_file.exists():
            return
        try:
            self._users_mtime = self.users_file.stat().st_mtime
            users = json.loads(self.users_file.read_text(encoding="utf-8"))
            user_id_to_code: dict[str, str] = {}
            full_name_to_code: dict[str, str] = {}
            for user_id, data in users.items():
                name = data.get("name", "")
                full_name = data.get("full_name", "")
                code = _extract_code(name)
                if code:
                    user_id_to_code[str(user_id)] = code
                    if full_name:
                        full_name_to_code[full_name.strip().lower()] = code
            self._user_id_to_code = user_id_to_code
            self._full_name_to_code = full_name_to_code
        except Exception as e:
            logger.error(f"Error loading user mappings: {e}")

    def _refresh_user_mappings_if_changed(self) -> None:
        try:
            if not self.users_file.exists():
                return
            mtime = self.users_file.stat().st_mtime
            if mtime != self._users_mtime:
                self._load_user_mappings()
        except Exception as e:
            logger.error(f"Error checking user mappings freshness: {e}")

    # ── Excel ────────────────────────────────────────────────────

    def list_months(self) -> list[str]:
        if not self.excel_path.exists():
            return []
        try:
            wb = load_workbook(self.excel_path, read_only=True, data_only=True)
            months = [s.upper() for s in wb.sheetnames if s.strip().upper() in MONTH_NAMES]
            wb.close()
            return months
        except Exception as e:
            logger.error(f"Error listing months: {e}")
            return []

    def _get_employees_from_excel(self, month: str) -> list[dict[str, Any]]:
        if not self.excel_path.exists():
            return []
        try:
            import pandas as pd
            wb = load_workbook(self.excel_path, data_only=True)
            sheet_name = next(
                (s for s in wb.sheetnames if s.strip().upper() == month.strip().upper()), None
            )
            if not sheet_name:
                wb.close()
                return []
            ws = wb[sheet_name]
            employees = []
            for row in range(3, 22):
                name = ws[f"A{row}"].value
                oklad = ws[f"AU{row}"].value
                excel_bonus = ws[f"BW{row}"].value
                if not name:
                    continue
                code = _extract_code(str(name))
                if not code:
                    continue
                employees.append({
                    "name": str(name).strip(),
                    "code": code,
                    "salary": float(oklad or 0),
                    "excel_bonus": float(excel_bonus or 0),
                })
            wb.close()

            # Read ОСН/ДОП/Цех from the main salary Excel (data.xlsx), matched by employee code
            try:
                from app.config import EXCEL_FILE
                df_main = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name, header=1, engine="openpyxl")
                if "ИМЯ" in df_main.columns:
                    extras_by_code: dict[str, dict] = {}
                    for _, r in df_main.iterrows():
                        row_name = str(r.get("ИМЯ") or "").strip()
                        if not row_name or row_name == "nan":
                            continue
                        code = (
                            self._full_name_to_code.get(row_name.lower())
                            or _extract_code(row_name)
                        )
                        if not code:
                            continue
                        extras_by_code[code] = {
                            "main_rate": float(r.get("ОСН", 0) or 0),
                            "main_shifts": float(r.get("ОСН.", 0) or 0),
                            "extra_rate": float(r.get("ДОП", 0) or 0),
                            "extra_shifts": float(r.get("ДОП.", 0) or 0),
                            "workshop": float(r.get("Цех", 0) or 0),
                        }
                    for emp in employees:
                        ex = extras_by_code.get(emp["code"], {})
                        emp["main_rate"] = ex.get("main_rate", 0.0)
                        emp["main_shifts"] = ex.get("main_shifts", 0.0)
                        emp["extra_rate"] = ex.get("extra_rate", 0.0)
                        emp["extra_shifts"] = ex.get("extra_shifts", 0.0)
                        emp["workshop"] = ex.get("workshop", 0.0)
            except Exception as e:
                logger.warning(f"Could not read rates/shifts from main Excel: {e}")

            return employees
        except Exception as e:
            logger.error(f"Error reading Excel sheet '{month}': {e}")
            return []

    # ── Advances ─────────────────────────────────────────────────

    def _load_advance_records(self) -> list[dict]:
        from app.data.payout_repository import PayoutRepository
        try:
            return PayoutRepository().load_all()
        except Exception as e:
            logger.error(f"Error reading advance requests from DB: {e}")
            return []

    def _resolve_code(self, row: dict) -> str | None:
        code = _extract_code(row.get("name"))
        if not code:
            code = self._user_id_to_code.get(str(row.get("user_id", "")))
        if not code:
            name_lower = str(row.get("name") or "").strip().lower()
            if name_lower:
                code = self._full_name_to_code.get(name_lower)
        return code

    def _get_advances_after_last_salary(self) -> dict[str, float]:
        """Original logic: advances since last salary payment (actual deduction amount)."""
        data = self._load_advance_records()
        VALID = {"Выплачено"}

        ops: dict[str, list[dict]] = {}
        for row in data:
            code = self._resolve_code(row)
            if code:
                ops.setdefault(code, []).append(row)

        out: dict[str, float] = {}
        for code, items in ops.items():
            items_sorted = sorted(items, key=lambda r: _parse_dt(_dt_field(r)))
            last_salary_dt = None
            for r in items_sorted:
                # A rejected (or still-pending) salary request never actually
                # paid out — it must not reset the "since last salary"
                # cutoff, or advances taken before it would silently drop
                # out of the deduction total.
                if r.get("payout_type") == "Зарплата" and r.get("status") in VALID:
                    last_salary_dt = _parse_dt(_dt_field(r))

            if last_salary_dt is None:
                out[code] = sum(
                    float(r.get("amount") or 0)
                    for r in items_sorted
                    if r.get("payout_type") == "Аванс" and r.get("status") in VALID
                )
                continue

            total = 0.0
            for r in items_sorted:
                if r.get("payout_type") != "Аванс" or r.get("status") not in VALID:
                    continue
                if _parse_dt(_dt_field(r)) <= last_salary_dt:
                    continue
                total += float(r.get("amount") or 0)
            out[code] = total

        return out

    def _get_advances_for_month(self, year: int, month_num: int) -> dict[str, float]:
        """Advances taken DURING a specific calendar month (for historical display)."""
        data = self._load_advance_records()
        VALID = {"Выплачено"}
        out: dict[str, float] = {}
        for row in data:
            if row.get("payout_type") != "Аванс" or row.get("status") not in VALID:
                continue
            dt = _parse_dt(_dt_field(row))
            if dt.year != year or dt.month != month_num:
                continue
            code = self._resolve_code(row)
            if code:
                out[code] = out.get(code, 0.0) + float(row.get("amount") or 0)
        return out

    def get_advances_history(self, month: str, year: int) -> dict[str, float]:
        """Public method: advances per employee for a specific month (for API)."""
        month_num = MONTH_NAMES.get(month.strip().upper())
        if not month_num:
            return {}
        return self._get_advances_for_month(year, month_num)

    # ── Bonuses / Penalties ───────────────────────────────────────

    def _get_bonuses_penalties_for_month(self, year: int, month: int):
        if not self.bonuses_penalties_file.exists():
            return {}, {}
        try:
            data = json.loads(self.bonuses_penalties_file.read_text(encoding="utf-8"))
        except Exception as e:
            logger.error(f"Error reading bonuses/penalties: {e}")
            return {}, {}

        bonuses: dict[str, float] = {}
        penalties: dict[str, float] = {}
        for r in data:
            code = self._user_id_to_code.get(str(r.get("employee_id", "")))
            if not code:
                code = self._full_name_to_code.get((r.get("name") or "").strip().lower())
            if not code:
                code = _extract_code(r.get("name"))
            if not code:
                continue
            dt = _parse_dt(r.get("date"))
            if dt.year != year or dt.month != month:
                continue
            amt = float(r.get("amount") or 0)
            if r.get("type") == "bonus":
                bonuses[code] = bonuses.get(code, 0.0) + amt
            elif r.get("type") == "penalty":
                penalties[code] = penalties.get(code, 0.0) + amt
        return bonuses, penalties

    # ── Sale transfers ────────────────────────────────────────────

    def _apply_sale_transfers(self, sales_data: dict, month_key: str) -> None:
        """Move order sales between employees and/or categories per manual
        corrections (hr.db).

        Firebird stays untouched; we only adjust the in-memory aggregates:
        subtract from the original employee/category and add to the new one.
        The destination category's commission rate applies from then on,
        since it's just a regular sale in that category once moved.

        All three categories (repair/cosmetics/shoes) carry both a scalar
        total and a per-order list (`{category}_orders`) — moving by doc_num
        keeps both in sync, so the salon-attribution report (which reads the
        order lists directly) stays correct after a manual correction too.
        """
        try:
            from app.services.sale_transfer_service import list_transfers
            transfers = list_transfers(month_key)
        except Exception as e:
            logger.error(f"Sale transfers error: {e}")
            return

        def _ensure(code: str) -> dict:
            entry = sales_data.get(code)
            if entry is None:
                entry = {
                    "repair": 0.0, "cosmetics": 0.0, "shoes": 0.0,
                    "repair_orders": [], "cosmetics_orders": [], "shoes_orders": [],
                }
                sales_data[code] = entry
            else:
                entry.setdefault("repair_orders", [])
                entry.setdefault("cosmetics_orders", [])
                entry.setdefault("shoes_orders", [])
            return entry

        for t in transfers:
            from_category = t.get("from_category") or t.get("category")
            to_category = t.get("to_category") or from_category
            from_code = t.get("from_code")
            to_code = t.get("to_code")
            amount = float(t.get("amount") or 0)
            doc_num = str(t.get("doc_num"))
            if not from_code or not to_code:
                continue
            if from_code == to_code and from_category == to_category:
                continue
            src = _ensure(from_code)
            dst = _ensure(to_code)

            src_key = f"{from_category}_orders"
            src_orders = src.get(src_key, []) or []
            # `shoes_orders` on the transfer itself carries the exact pairs
            # being moved when a doc_num covers more than one pair and only
            # some of them should move; falls back to matching by doc_num.
            moved_orders = t.get("shoes_orders") or [
                o for o in src_orders if str(o.get("doc_num")) == doc_num
            ]
            if not moved_orders:
                # No per-order rows matched (e.g. a doc_num typo, or data
                # from before per-order tracking existed) — fall back to a
                # synthetic single order so the scalar total still moves.
                moved_orders = [{"doc_num": doc_num, "kredit": amount}]
            src[src_key] = [o for o in src_orders if str(o.get("doc_num")) != doc_num]
            src[from_category] = sum(o.get("kredit", 0.0) for o in src[src_key])

            dst_key = f"{to_category}_orders"
            dst.setdefault(dst_key, []).extend(moved_orders)
            dst[to_category] = sum(o.get("kredit", 0.0) for o in dst.get(dst_key, []))

    # ── Commission ────────────────────────────────────────────────

    def _commission(self, sales, plan, rate_hi, rate_lo):
        if plan and plan > 0:
            fulfillment = sales / plan
            rate = rate_hi if fulfillment >= PLAN_THRESHOLD else rate_lo
        else:
            fulfillment = 0.0
            rate = rate_lo
        return fulfillment, rate, sales * rate

    # ── Main calculation ──────────────────────────────────────────

    async def calculate_payroll(self, month: str, year: int | None = None) -> tuple[list[PayrollRow], list[str]]:
        """Returns (rows, unknown_location_codes)."""
        rows, unknown_codes, _order_detail = await self._calculate_payroll_internal(month, year)
        return rows, unknown_codes

    async def _calculate_payroll_internal(
        self, month: str, year: int | None = None
    ) -> tuple[list[PayrollRow], list[str], dict[str, dict]]:
        """Same computation as `calculate_payroll`, plus per-employee raw order
        detail (`order_detail`) needed by the by-salon report — kept as a
        single internal worker so both callers share one Excel parse, one
        Firebird round-trip, and one sale-transfer application."""
        self._refresh_user_mappings_if_changed()
        month = month.strip().upper()
        month_num = MONTH_NAMES.get(month)
        if not month_num:
            return [], [], {}

        if year is None:
            year = datetime.now().year

        month_key = make_month_key(month, year)

        # Excel parsing and the Firebird round-trips below are blocking calls;
        # run them off the event loop so a slow report (e.g. a multi-month
        # summary firing several of these concurrently) doesn't freeze every
        # other request the API process is serving.
        employees = await asyncio.to_thread(self._get_employees_from_excel, month)
        if not employees:
            return [], [], {}

        try:
            sales_data = await asyncio.to_thread(self.firebird.get_all_sales, year, month_num)
        except Exception as e:
            logger.error(f"Firebird error: {e}")
            sales_data = {}

        # Apply manual sale transfers (corrections layered on top of Firebird).
        self._apply_sale_transfers(sales_data, month_key)

        advances_map = self._get_advances_after_last_salary()
        advances_month_map = self._get_advances_for_month(year, month_num)
        bonuses_map, penalties_map = self._get_bonuses_penalties_for_month(year, month_num)
        plans_map = self.plans_repo.get_plans_map(month_key=month_key)
        settlements_map = self.settlement_repo.get_settlements_map(month_key)

        # ── Schedule-based shift counting ─────────────────────────
        schedule = await asyncio.to_thread(_parse_schedule_from_excel, month, year)
        days_in_month = calendar.monthrange(year, month_num)[1]
        loc_plans_map = self.location_repo.plans_map(month_key)

        # Build lookup: code → shifts_by_point from schedule
        # Schedule rows use full names (e.g. "Катя 2201"), extract code to match
        schedule_by_code: dict[str, dict[str, int]] = {}
        for emp_name, shifts in schedule.items():
            code = _extract_code(emp_name)
            if code:
                schedule_by_code[code] = shifts

        # Detect unknown location codes (present in schedule but not in location repo)
        known_codes = set(self.location_repo.codes_dict().keys())
        unknown_codes: list[str] = sorted({
            pt_code
            for shifts in schedule_by_code.values()
            for pt_code in shifts
            if pt_code not in known_codes
        })

        results = []
        order_detail: dict[str, dict] = {}
        for emp in employees:
            code = emp["code"]
            name = emp["name"]
            excel_bonus = emp.get("excel_bonus", 0.0)

            main_rate   = emp.get("main_rate", 0.0)
            extra_rate  = emp.get("extra_rate", 0.0)

            # Shift counts from schedule (prefer live count; fall back to Excel columns)
            shifts_by_point = schedule_by_code.get(code, {})
            if shifts_by_point:
                total_shifts  = sum(shifts_by_point.values())
                main_shifts   = min(total_shifts, 15)
                extra_shifts  = max(0, total_shifts - 15)
                base_salary   = _calc_base_salary(total_shifts, main_rate, extra_rate)
            else:
                # Fallback: use values already parsed from data.xlsx
                main_shifts  = int(emp.get("main_shifts", 0))
                extra_shifts = int(emp.get("extra_shifts", 0))
                total_shifts = main_shifts + extra_shifts
                if main_rate or extra_rate:
                    base_salary = _calc_base_salary(total_shifts, main_rate, extra_rate)
                else:
                    base_salary = emp.get("salary", 0.0)

            emp_sales = sales_data.get(code, {})
            repair_sales = emp_sales.get("repair", 0.0)
            cosmetics_sales = emp_sales.get("cosmetics", 0.0)
            shoes_sales = emp_sales.get("shoes", 0.0)
            shoes_order_items = emp_sales.get("shoes_orders", [])
            shoes_orders = [o["doc_num"] for o in shoes_order_items if isinstance(o, dict)]
            repair_order_items = emp_sales.get("repair_orders", [])
            cosmetics_order_items = emp_sales.get("cosmetics_orders", [])

            plan = plans_map.get(code)
            ignore_kpi = plan.ignore_kpi if plan else False

            # ── Individual plan: manual override or auto from location plans ──
            if plan and (plan.repair_plan or plan.cosmetics_plan or plan.shoes_plan):
                repair_plan    = plan.repair_plan
                cosmetics_plan = plan.cosmetics_plan
                shoes_plan     = plan.shoes_plan
            elif shifts_by_point and loc_plans_map:
                # Auto-calculate: Σ (daily_location_plan × days_at_location)
                repair_plan = cosmetics_plan = shoes_plan = 0.0
                for pt_code, pt_days in shifts_by_point.items():
                    lp = loc_plans_map.get(pt_code)
                    if lp and days_in_month > 0:
                        repair_plan    += lp.repair_plan    / days_in_month * pt_days
                        cosmetics_plan += lp.cosmetics_plan / days_in_month * pt_days
                        shoes_plan     += lp.shoes_plan     / days_in_month * pt_days
                repair_plan    = round(repair_plan)
                cosmetics_plan = round(cosmetics_plan)
                shoes_plan     = round(shoes_plan)
            else:
                repair_plan    = plan.repair_plan    if plan else 0.0
                cosmetics_plan = plan.cosmetics_plan if plan else 0.0
                shoes_plan     = plan.shoes_plan     if plan else 0.0
            force_max = plan.force_max if plan else []
            force_min = plan.force_min if plan else []

            # Repair
            if "repair" in force_max:
                repair_fulfillment, repair_rate = 1.0, REPAIR_RATE_HIGH
                repair_commission = repair_sales * REPAIR_RATE_HIGH
            elif "repair" in force_min:
                repair_fulfillment, repair_rate = 0.0, REPAIR_RATE_LOW
                repair_commission = repair_sales * REPAIR_RATE_LOW
            else:
                repair_fulfillment, repair_rate, repair_commission = self._commission(
                    repair_sales, repair_plan, REPAIR_RATE_HIGH, REPAIR_RATE_LOW
                )

            # Cosmetics
            if "cosmetics" in force_max:
                cosmetics_fulfillment, cosmetics_rate = 1.0, COSMETICS_RATE_HIGH
                cosmetics_commission = cosmetics_sales * COSMETICS_RATE_HIGH
            elif "cosmetics" in force_min:
                cosmetics_fulfillment, cosmetics_rate = 0.0, COSMETICS_RATE_LOW
                cosmetics_commission = cosmetics_sales * COSMETICS_RATE_LOW
            else:
                cosmetics_fulfillment, cosmetics_rate, cosmetics_commission = self._commission(
                    cosmetics_sales, cosmetics_plan, COSMETICS_RATE_HIGH, COSMETICS_RATE_LOW
                )

            shoes_fulfillment = 0.0
            shoes_rate = 0.0

            if ignore_kpi:
                repair_commission = cosmetics_commission = shoes_commission = 0.0
            else:
                if "shoes" in force_max:
                    shoes_commission = 1000.0 * sum(1 for o in shoes_order_items if isinstance(o, dict))
                elif "shoes" in force_min:
                    shoes_commission = 500.0 * sum(1 for o in shoes_order_items if isinstance(o, dict))
                else:
                    shoes_commission = sum(
                        1000 if o["kredit"] > 11000 else 500
                        for o in shoes_order_items if isinstance(o, dict)
                    )

            bonuses = bonuses_map.get(code, 0.0)
            penalties = penalties_map.get(code, 0.0)
            advances = advances_map.get(code, 0.0)
            advances_this_month = advances_month_map.get(code, 0.0)
            settlement_paid = settlements_map.get(code, False)

            total_commission = repair_commission + cosmetics_commission + shoes_commission
            total_gross = base_salary + total_commission + bonuses + excel_bonus
            total_deductions = advances + penalties
            total_net = total_gross - total_deductions

            results.append(PayrollRow(
                employee_code=code,
                employee_name=name,
                base_salary=base_salary,
                repair_sales=repair_sales,
                cosmetics_sales=cosmetics_sales,
                shoes_sales=shoes_sales,
                repair_plan=repair_plan,
                cosmetics_plan=cosmetics_plan,
                shoes_plan=shoes_plan,
                repair_fulfillment=repair_fulfillment,
                cosmetics_fulfillment=cosmetics_fulfillment,
                shoes_fulfillment=shoes_fulfillment,
                repair_rate=repair_rate,
                cosmetics_rate=cosmetics_rate,
                shoes_rate=shoes_rate,
                repair_commission=repair_commission,
                cosmetics_commission=cosmetics_commission,
                shoes_commission=shoes_commission,
                bonuses=bonuses,
                excel_bonus=excel_bonus,
                penalties=penalties,
                advances=advances,
                advances_this_month=advances_this_month,
                ignore_kpi=ignore_kpi,
                force_max=force_max,
                force_min=force_min,
                shoes_orders=shoes_orders,
                total_commission=total_commission,
                total_gross=total_gross,
                total_deductions=total_deductions,
                total_net=total_net,
                settlement_paid=settlement_paid,
                main_rate=main_rate,
                main_shifts=main_shifts,
                extra_rate=extra_rate,
                extra_shifts=extra_shifts,
                workshop_commission=emp.get("workshop", 0.0),
                shifts_by_point=shifts_by_point,
            ))

            order_detail[code] = {
                "repair_orders": repair_order_items,
                "cosmetics_orders": cosmetics_order_items,
                "shoes_order_items": shoes_order_items,
            }

        return results, unknown_codes, order_detail

    async def get_employee_details(self, employee_code: str, month: str, year: int | None = None):
        rows, _ = await self.calculate_payroll(month, year)
        for row in rows:
            if row.employee_code == employee_code:
                return row
        return None

    UNALLOCATED_SALON_ID = "unallocated"
    UNALLOCATED_SALON_NAME = "Не определено"

    async def get_payroll_by_salon(self, month: str, year: int | None = None) -> dict:
        """Same commission/oklad totals as `calculate_payroll`, split by the
        salon each order/shift happened at. An employee's monthly rate
        (already resolved for plan fulfillment / force_max / force_min) is
        applied to that employee's sales at each specific salon — no
        per-salon plan is computed or invented."""
        rows, unknown_codes, order_detail = await self._calculate_payroll_internal(month, year)
        month = month.strip().upper()
        if year is None:
            year = datetime.now().year
        month_key = make_month_key(month, year)

        UNALLOC_ID = self.UNALLOCATED_SALON_ID
        UNALLOC_NAME = self.UNALLOCATED_SALON_NAME

        salons: dict[str, dict] = {}

        def _bucket(salon_id: str, salon_name: str) -> dict:
            b = salons.get(salon_id)
            if b is None:
                b = {
                    "salon_id": salon_id,
                    "salon_name": salon_name,
                    "oklad": 0.0,
                    "bonuses": 0.0,
                    "repair_commission": 0.0,
                    "cosmetics_commission": 0.0,
                    "shoes_commission": 0.0,
                    "total": 0.0,
                    "employees": {},
                }
                salons[salon_id] = b
            return b

        def _emp_entry(bucket: dict, code: str, name: str) -> dict:
            e = bucket["employees"].get(code)
            if e is None:
                e = {
                    "employee_code": code,
                    "employee_name": name,
                    "oklad": 0.0,
                    "bonuses": 0.0,
                    "repair_commission": 0.0,
                    "cosmetics_commission": 0.0,
                    "shoes_commission": 0.0,
                    "total": 0.0,
                }
                bucket["employees"][code] = e
            return e

        def _add(bucket: dict, emp: dict, field: str, amount: float) -> None:
            if not amount:
                return
            bucket[field] += amount
            bucket["total"] += amount
            emp[field] += amount
            emp["total"] += amount

        grand_total = {
            "oklad": 0.0,
            "bonuses": 0.0,
            "repair_commission": 0.0,
            "cosmetics_commission": 0.0,
            "shoes_commission": 0.0,
            "total": 0.0,
        }

        def _resolve_by_order(doc_num: str | None) -> tuple[str, str]:
            salon = self.salon_repo.get_by_order_code(_order_salon_code(doc_num))
            if salon:
                return salon.id, salon.name
            return UNALLOC_ID, UNALLOC_NAME

        for row in rows:
            code = row.employee_code
            name = row.employee_name

            # ── Employee-level amounts (oklad, премии): not tied to any
            # specific sale, so split proportionally across the salons
            # worked this month — same rule as the oklad itself.
            shifts_by_point = row.shifts_by_point or {}
            total_shifts = sum(shifts_by_point.values())

            def _split_by_shifts(field: str, amount: float) -> None:
                if not amount:
                    return
                if total_shifts > 0:
                    for pt_code, pt_shifts in shifts_by_point.items():
                        if not pt_shifts:
                            continue
                        salon = self.salon_repo.get_by_code(pt_code)
                        salon_id = salon.id if salon else UNALLOC_ID
                        salon_name = salon.name if salon else UNALLOC_NAME
                        bucket = _bucket(salon_id, salon_name)
                        emp = _emp_entry(bucket, code, name)
                        portion = amount * pt_shifts / total_shifts
                        _add(bucket, emp, field, portion)
                        grand_total[field] += portion
                else:
                    bucket = _bucket(UNALLOC_ID, UNALLOC_NAME)
                    emp = _emp_entry(bucket, code, name)
                    _add(bucket, emp, field, amount)
                    grand_total[field] += amount

            _split_by_shifts("oklad", row.base_salary)
            _split_by_shifts("bonuses", (row.bonuses or 0.0) + (row.excel_bonus or 0.0))

            # ignore_kpi zeroes the row's final commission but NOT its
            # resolved rates — multiplying rate × per-order sales here would
            # silently resurrect commission the main calculation zeroed out.
            if row.ignore_kpi:
                continue

            detail = order_detail.get(code, {})

            for order in detail.get("repair_orders", []):
                if not isinstance(order, dict):
                    continue
                kredit = float(order.get("kredit") or 0.0)
                if not kredit:
                    continue
                commission = kredit * row.repair_rate
                salon_id, salon_name = _resolve_by_order(order.get("doc_num"))
                bucket = _bucket(salon_id, salon_name)
                emp = _emp_entry(bucket, code, name)
                _add(bucket, emp, "repair_commission", commission)
                grand_total["repair_commission"] += commission

            for order in detail.get("cosmetics_orders", []):
                if not isinstance(order, dict):
                    continue
                kredit = float(order.get("kredit") or 0.0)
                if not kredit:
                    continue
                commission = kredit * row.cosmetics_rate
                salon_id, salon_name = _resolve_by_order(order.get("doc_num"))
                bucket = _bucket(salon_id, salon_name)
                emp = _emp_entry(bucket, code, name)
                _add(bucket, emp, "cosmetics_commission", commission)
                grand_total["cosmetics_commission"] += commission

            for order in detail.get("shoes_order_items", []):
                if not isinstance(order, dict):
                    continue
                kredit = float(order.get("kredit") or 0.0)
                if "shoes" in row.force_max:
                    commission = 1000.0
                elif "shoes" in row.force_min:
                    commission = 500.0
                else:
                    commission = 1000.0 if kredit > 11000 else 500.0
                salon_id, salon_name = _resolve_by_order(order.get("doc_num"))
                bucket = _bucket(salon_id, salon_name)
                emp = _emp_entry(bucket, code, name)
                _add(bucket, emp, "shoes_commission", commission)
                grand_total["shoes_commission"] += commission

        grand_total["total"] = (
            grand_total["oklad"]
            + grand_total["bonuses"]
            + grand_total["repair_commission"]
            + grand_total["cosmetics_commission"]
            + grand_total["shoes_commission"]
        )

        salon_list = []
        for bucket in salons.values():
            bucket = dict(bucket)
            bucket["employees"] = sorted(
                bucket["employees"].values(), key=lambda e: e["employee_name"]
            )
            salon_list.append(bucket)
        salon_list.sort(key=lambda s: (s["salon_id"] == UNALLOC_ID, s["salon_name"]))

        return {
            "month_key": month_key,
            "salons": salon_list,
            "unknown_codes": unknown_codes,
            "grand_total": grand_total,
        }

    def get_code_for_employee(self, employee_id: str | None = None, full_name: str | None = None) -> str | None:
        """Resolve payroll employee_code by employee_id (telegram id) or full_name."""
        self._refresh_user_mappings_if_changed()
        if employee_id:
            code = self._user_id_to_code.get(str(employee_id))
            if code:
                return code
        if full_name:
            code = self._full_name_to_code.get(full_name.strip().lower())
            if code:
                return code
            # Fallback: extract 4-digit code from display name (e.g. "Иванов И.И. 1234")
            code = _extract_code(full_name)
            if code:
                return code
        return None


_payroll_service: PayrollService | None = None


def get_payroll_service() -> PayrollService:
    global _payroll_service
    if _payroll_service is None:
        _payroll_service = PayrollService()
    return _payroll_service
