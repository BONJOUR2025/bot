from __future__ import annotations

import json
import os
import re
import textwrap
from datetime import datetime

import pandas as pd
from fpdf import FPDF
from openpyxl import load_workbook

from ..config import ADVANCE_REQUESTS_FILE, EXCEL_FILE
from ..utils.logger import log


def unmerge_cells(sheet):
    """Разъединяет объединённые ячейки и копирует их значение во все ячейки диапазона."""
    merged_ranges = list(sheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        sheet.unmerge_cells(str(merged_range))
        top_left_value = sheet.cell(
            row=merged_range.min_row, column=merged_range.min_col
        ).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                sheet.cell(row=row, column=col, value=top_left_value)
    return sheet


def get_cell_comment(sheet_name, row_index, column_letter):
    """Получает примечание из указанной ячейки Excel."""
    if not os.path.exists(EXCEL_FILE):
        log(f"❌ Error: File {EXCEL_FILE} not found!")
        return "File error"
    try:
        # read_only=False, data_only=False to get comments
        workbook = load_workbook(EXCEL_FILE, read_only=False, data_only=False)
        if sheet_name not in workbook.sheetnames:
            log(f"❌ Error: Sheet {sheet_name} not found!")
            workbook.close()
            return "Sheet error"
        sheet = workbook[sheet_name]
        # row_index is pandas index (0-based from row 3 in Excel)
        # Excel row = row_index + 3 (header on row 2, data starts row 3)
        cell_ref = f"{column_letter}{row_index + 3}"
        cell = sheet[cell_ref]
        comment_text = cell.comment.text.strip() if cell.comment else "No comment"
        workbook.close()
        return comment_text
    except Exception as e:
        log(f"❌ Error loading comment from {column_letter}{row_index + 3}: {e}")
        return "Error"


def load_data(sheet_name=None):
    """
    Загружает данные из Excel.
    :param sheet_name: Название листа (месяц) или None для получения списка листов.
    :return: DataFrame с данными или список листов.
    """
    if not os.path.exists(EXCEL_FILE):
        log(f"❌ Ошибка: Файл Excel не найден по пути {EXCEL_FILE}")
        return None

    try:
        # Read fresh each time - no caching
        if sheet_name is None:
            # Just get sheet names
            xls = pd.ExcelFile(EXCEL_FILE, engine='openpyxl')
            sheets = xls.sheet_names
            xls.close()
            log(f"📂 Доступные листы в файле: {sheets}")
            return sheets

        # Read specific sheet directly (no ExcelFile caching)
        log(f"📖 Читаю лист '{sheet_name}' из {EXCEL_FILE}")
        df = pd.read_excel(
            EXCEL_FILE,
            sheet_name=sheet_name,
            header=1,
            engine='openpyxl'
        )
        log(f"✅ Загружено {len(df)} строк, колонки: {list(df.columns)[:10]}...")
        return df
    except Exception as e:
        log(f"❌ Ошибка при загрузке Excel: {e}")
        return None


def export_to_pdf(sheet_name="ЯНВАРЬ"):
    """Экспортирует данные в PDF."""
    try:
        from fpdf import FPDF

        data = load_data(sheet_name)
        if data is None:
            return None
        filename = f"data_{sheet_name}.pdf"
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", size=10)
        pdf.cell(200, 10, f"Data for {sheet_name}", ln=True, align="C")
        for index, row in data.iterrows():
            row_text = " | ".join(str(x) for x in row)
            pdf.cell(200, 10, row_text, ln=True, align="L")
        pdf.output(filename)
        return filename
    except Exception as e:
        log(f"Error exporting to PDF: {e}")
        return None


def clean_line(text: str) -> str:
    return re.sub(r"[^\x00-\x7Fа-яА-ЯёЁ0-9\s.,!?@\-:;()|₽💳🏠✅❌]+", "", text)


def export_advances_to_pdf(
    filter_type=None,
    status=None,
    name=None,
    method=None,
    after_date=None,
    before_date=None,
    filename="advance_report.pdf",
):
    try:
        with open(ADVANCE_REQUESTS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        log(f"❌ Ошибка чтения файла: {e}")
        return None

    if not data:
        log("⚠️ Нет данных для экспорта.")
        return None

    # Фильтрация по параметрам
    def match_filters(entry):
        if filter_type and entry.get("payout_type") != filter_type:
            return False
        if status and entry.get("status") != status:
            return False
        if name and name.lower() not in str(entry.get("name", "")).lower():
            return False
        if method and entry.get("method") != method:
            return False
        if after_date:
            ts = entry.get("timestamp")
            if ts:
                try:
                    dt = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
                    if dt < datetime.strptime(after_date, "%Y-%m-%d"):
                        return False
                except Exception:
                    return False
        if before_date:
            ts = entry.get("timestamp")
            if ts:
                try:
                    dt = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
                    if dt > datetime.strptime(before_date, "%Y-%m-%d"):
                        return False
                except Exception:
                    return False
        return True

    data = [d for d in data if match_filters(d)]

    from ..config import FONT_PATH

    font_path = FONT_PATH
    bold_font = FONT_PATH.replace(".ttf", "-Bold.ttf")
    pdf = FPDF()
    pdf.add_page()
    if os.path.exists(font_path):
        pdf.add_font("DejaVu", "", font_path, uni=True)
        if os.path.exists(bold_font):
            pdf.add_font("DejaVu", "B", bold_font, uni=True)
        pdf.set_font("DejaVu", "", 10)
    else:
        log(
            f"⚠️ Шрифт не найден: {font_path}. Используется стандартный Arial"
        )
        pdf.set_font("Arial", size=10)
    pdf.set_auto_page_break(auto=True, margin=15)

    pdf.cell(200, 10, "📄 Отчёт по выплатам", ln=True, align="C")

    for idx, r in enumerate(data, 1):
        try:
            timestamp = str(r.get("timestamp", "—"))
            name_val = str(r.get("name", "—"))
            amount = str(r.get("amount", 0))
            method = str(r.get("method", "—"))
            payout_type = str(r.get("payout_type", "—"))
            status_val = str(r.get("status", "—"))

            line = f"{idx}) {timestamp} | {name_val} | {amount} ₽ | {method} | {payout_type} | {status_val}"
            line = clean_line(line)

            if len(line) > 1000:
                line = line[:1000] + "..."

            for chunk in textwrap.wrap(line, width=110):
                pdf.cell(0, 8, txt=chunk, ln=True)
        except Exception as e:
            log(f"❌ Ошибка в строке {idx}: {e}")
            continue

    try:
        pdf.output(filename)
        log(f"✅ PDF отчёт сохранён: {filename}")
        return filename
    except Exception as e:
        log(f"❌ Ошибка сохранения PDF: {e}")
        return None
