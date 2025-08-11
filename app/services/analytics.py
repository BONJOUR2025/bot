from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional, Any, List
import asyncio
import os

import pandas as pd

from ..config import (
    EXCEL_FILE,
    SALES_FILE,
    FIREBIRD_DB,
    FIREBIRD_USER,
    FIREBIRD_PASSWORD,
)
from openpyxl import load_workbook
from .firebird_service import FirebirdService
from ..core.constants import MONTHS_RU
from ..utils.logger import log
import re

# Mapping of 4 digit codes to employee names used in sales analytics
EMPLOYEE_CODE_MAP = {
    "0102": "Вера 0102",
    "2602": "Анастасия 2602",
    "7272": "Арина 7272",
    "1505": "Александр 1505",
    "2404": "Эмиль 2404",
    "5984": "Полина 5984",
    "0704": "Наталья 0704",
    "2201": "Катя 2201",
    "1606": "Лали 1606",
    "0104": "Екатерина 0104",
    "2006": "Ирина 2006",
    "1802": "Полина 1802",
    "1996": "Вероника 1996",
    "2405": "Ира 2405",
    "3007": "Юля 3007",
    "2104": "Алекс 2104",
    "0208": "Марина 0208",
}


def _count_shifts_from_excel(months: list[str]) -> dict[str, int]:
    """Return mapping of employee name to shift count for the given months."""
    counts: dict[str, int] = {}
    if not os.path.exists(EXCEL_FILE):
        return counts
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
    except Exception as exc:
        log(f"❌ Failed to read Excel for shifts: {exc}")
        return counts

    CODES = {"Ц", "Ох", "М", "А", "Оз", "П", "Р"}

    for month in months:
        sheet = None
        for title in wb.sheetnames:
            if title.upper().startswith(month.upper()):
                sheet = wb[title]
                break
        if sheet is None:
            continue

        start_col = None
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col).value
            if isinstance(val, (int, float)) and int(val) == 1:
                start_col = col
                break
        if start_col is None:
            continue

        day_cols: list[int] = []
        col = start_col
        while col <= sheet.max_column:
            val = sheet.cell(row=1, column=col).value
            if isinstance(val, (int, float)) and 1 <= int(val) <= 31:
                day_cols.append(col)
                col += 1
            else:
                break

        for row in range(3, sheet.max_row + 1):
            name_cell = sheet.cell(row=row, column=1).value
            if not name_cell:
                continue
            name = map_employee_by_code(str(name_cell).strip())
            cnt = 0
            for c in day_cols:
                val = sheet.cell(row=row, column=c).value
                if str(val).strip() in CODES:
                    cnt += 1
            counts[name] = counts.get(name, 0) + cnt

    return counts


def map_employee_by_code(description: str | None) -> str:
    """Return employee name by the last 4 digits in the description."""
    if not description:
        return description or ""
    match = re.search(r"(\d{4})\s*$", str(description))
    if match:
        code = match.group(1)
        return EMPLOYEE_CODE_MAP.get(code, str(description))
    return str(description)


class AnalyticsService:
    """Load sales analytics from the salary Excel workbook."""

    def __init__(self, fb_service: FirebirdService | None = None) -> None:
        self._data: Optional[dict] = None
        self._updated_at: Optional[datetime] = None
        self._details_df: Optional[pd.DataFrame] = None
        self._details_mtime: float = 0.0
        if fb_service:
            self._fb = fb_service
        elif FIREBIRD_DB and FIREBIRD_USER and FIREBIRD_PASSWORD:
            self._fb = FirebirdService(FIREBIRD_DB, FIREBIRD_USER, FIREBIRD_PASSWORD)
        else:
            self._fb = None

    def _collect_sales(self) -> dict:
        try:
            xls = pd.ExcelFile(EXCEL_FILE)
        except Exception as exc:
            log(f"❌ Failed to read Excel: {exc}")
            return {
                "repair_sum": 0,
                "repair_count": 0,
                "cosmetics_sum": 0,
                "cosmetics_count": 0,
                "updated_at": datetime.utcnow().isoformat(),
            }

        repair_sum = repair_count = 0
        cosmetics_sum = cosmetics_count = 0
        months = [m.upper() for m in MONTHS_RU]
        for sheet in xls.sheet_names:
            name = sheet.split()[0].upper()
            if name not in months:
                continue
            try:
                df = pd.read_excel(xls, sheet_name=sheet, header=1)
            except Exception:
                continue
            cols = {str(c).strip(): c for c in df.columns}
            if "Ремонт" in cols:
                col = cols["Ремонт"]
                series = pd.to_numeric(df[col], errors="coerce").fillna(0)
                repair_sum += series.sum()
                repair_count += (series != 0).sum()
            if "Косметика" in cols:
                col = cols["Косметика"]
                series = pd.to_numeric(df[col], errors="coerce").fillna(0)
                cosmetics_sum += series.sum()
                cosmetics_count += (series != 0).sum()

        self._updated_at = datetime.utcnow()
        data = {
            "repair_sum": int(repair_sum),
            "repair_count": int(repair_count),
            "cosmetics_sum": int(cosmetics_sum),
            "cosmetics_count": int(cosmetics_count),
            "updated_at": self._updated_at.isoformat(),
        }
        self._data = data
        return data

    async def get_sales(self) -> dict:
        if self._data is None:
            return self._collect_sales()
        return self._data

    async def refresh_sales(self) -> dict:
        return self._collect_sales()

    async def refresh_sales_details(self) -> None:
        """Force refresh cached sales details."""
        await self._ensure_details_df(force=True)

    async def _read_sales_file(self) -> pd.DataFrame:
        def _read() -> pd.DataFrame:
            return pd.read_excel(
                SALES_FILE,
                header=None,
                usecols="A,B,E,G,I",
                names=["period", "order_number", "employee", "item", "cost"],
            )

        return await asyncio.to_thread(_read)

    async def _ensure_details_df(self, force: bool = False) -> pd.DataFrame | None:
        try:
            mtime = os.path.getmtime(SALES_FILE)
        except Exception as exc:
            log(f"❌ Failed to stat sales file: {exc}")
            return None

        if self._details_df is not None and not force and mtime == self._details_mtime:
            return self._details_df

        start = datetime.utcnow()
        try:
            df = await self._read_sales_file()
        except Exception as exc:
            log(f"❌ Failed to read sales details: {exc}")
            return None

        df.columns = ["period", "order_number", "employee", "item", "cost"]
        df.dropna(how="all", inplace=True)
        df["cost"] = pd.to_numeric(df["cost"], errors="coerce").fillna(0)
        df["order_number"] = df["order_number"].astype(str).str.strip()
        df = df[
            df["order_number"].notna()
            & (df["order_number"] != "")
            & (df["order_number"].str.lower() != "nan")
        ]
        df["period"] = pd.to_datetime(df["period"], errors="coerce", dayfirst=True)
        dropped = df["period"].isna().sum()
        if dropped:
            log(f"⚠️ Dropped {dropped} rows with invalid period")
        df = df.dropna(subset=["period"]).reset_index(drop=True)

        # Replace employee field with mapped name if possible
        df["employee"] = df["employee"].apply(map_employee_by_code)

        self._details_df = df
        self._details_mtime = mtime
        elapsed = (datetime.utcnow() - start).total_seconds()
        log(f"✅ Loaded {len(df)} sales details in {elapsed:.2f}s")
        return df

    async def _firebird_details(
        self,
        date_from: str | None,
        date_to: str | None,
        creater_ids: list[str] | None,
        user_ids: list[str] | None,
        folder_ids: list[str],
        code_substr: str | None,
        name_substr: str | None,
        min_kredit: float | None,
        max_kredit: float | None,
        doc_num_substr: str | None,
        item_type: str | None,
        page: int,
        page_size: int,
    ) -> dict | None:
        if not self._fb:
            return None

        filters = ["docs_order_history.status_id = 5"]
        params: list[Any] = []
        if date_from:
            filters.append("docs.doc_date >= ?")
            params.append(date_from)
        if date_to:
            filters.append("docs.doc_date <= ?")
            params.append(date_to)
        if creater_ids:
            placeholders = ",".join(["?"] * len(creater_ids))
            filters.append(f"docs_order.creater_id IN ({placeholders})")
            params.extend(creater_ids)
        if user_ids:
            placeholders = ",".join(["?"] * len(user_ids))
            filters.append(f"users.user_id IN ({placeholders})")
            params.extend(user_ids)
        if folder_ids:
            placeholders = ",".join(["?"] * len(folder_ids))
            filters.append(f"tovars_tbl.folder_id IN ({placeholders})")
            params.extend(folder_ids)
        if code_substr:
            filters.append("tovars_tbl.code CONTAINING ?")
            params.append(code_substr)
        if name_substr:
            filters.append("tovars_tbl.name CONTAINING ?")
            params.append(name_substr)
        if min_kredit is not None:
            filters.append("doc_order_lines.kredit >= ?")
            params.append(min_kredit)
        if max_kredit is not None:
            filters.append("doc_order_lines.kredit <= ?")
            params.append(max_kredit)
        if doc_num_substr:
            filters.append("docs.doc_num CONTAINING ?")
            params.append(doc_num_substr)

        where_clause = " AND ".join(filters)
        if where_clause:
            where_clause = "WHERE " + where_clause

        lines_from = (
            "FROM doc_order_lines "
            "INNER JOIN docs_order ON doc_order_lines.doc_order_id = docs_order.id "
            "INNER JOIN docs_order_history ON docs_order.id = docs_order_history.doc_order_id "
            "INNER JOIN docs ON docs_order.doc_id = docs.doc_id "
            "INNER JOIN contragents ON docs.contragent_id = contragents.contr_id "
            "INNER JOIN tovars_tbl ON doc_order_lines.tovar_id = tovars_tbl.tovar_id "
            "INNER JOIN users ON docs_order.creater_id = users.user_id "
        )

        services_from = (
            "FROM doc_order_services "
            "INNER JOIN docs_order ON doc_order_services.doc_order_id = docs_order.id "
            "INNER JOIN docs_order_history ON docs_order.id = docs_order_history.doc_order_id "
            "INNER JOIN docs ON docs_order.doc_id = docs.doc_id "
            "INNER JOIN contragents ON docs.contragent_id = contragents.contr_id "
            "INNER JOIN tovars_tbl ON doc_order_services.tovar_id = tovars_tbl.tovar_id "
            "INNER JOIN users ON docs_order.creater_id = users.user_id "
        )

        lines_where = f"{where_clause} {'AND' if where_clause else 'WHERE'} doc_order_lines.kredit <> 0"
        services_where = f"{where_clause} {'AND' if where_clause else 'WHERE'} doc_order_services.kredit <> 0"

        lines_select = (
            "SELECT docs.doc_date AS doc_date, docs.doc_num AS doc_number, "
            "docs_order.creater_id AS creator_id, users.user_id AS user_id, "
            "users.description AS description, tovars_tbl.code AS item_code, "
            "tovars_tbl.name AS item_name, doc_order_lines.kredit AS kredit, 'cosmetics' AS item_type "
            f"{lines_from} {lines_where}"
        )

        services_select = (
            "SELECT docs.doc_date AS doc_date, docs.doc_num AS doc_number, "
            "docs_order.creater_id AS creator_id, users.user_id AS user_id, "
            "users.description AS description, tovars_tbl.code AS item_code, "
            "tovars_tbl.name AS item_name, doc_order_services.kredit AS kredit, 'service' AS item_type "
            f"{services_from} {services_where}"
        )

        start_row = max(1, (page - 1) * page_size + 1)
        end_row = page * page_size

        if item_type == 'cosmetics':
            final_query = f"{lines_select} ORDER BY doc_date, doc_number ROWS ? TO ?"
            final_count_query = f"SELECT COUNT(*) as cnt {lines_from} {lines_where}"
            query_params = params + [start_row, end_row]
            count_params = params
            cache_key = ("cosmetics", tuple(params), page, page_size)
        elif item_type == 'services':
            final_query = f"{services_select} ORDER BY doc_date, doc_number ROWS ? TO ?"
            final_count_query = f"SELECT COUNT(*) as cnt {services_from} {services_where}"
            query_params = params + [start_row, end_row]
            count_params = params
            cache_key = ("services", tuple(params), page, page_size)
        else:
            union_select = f"{lines_select} UNION ALL {services_select}"
            final_query = f"SELECT * FROM ({union_select}) ORDER BY doc_date, doc_number ROWS ? TO ?"
            final_count_query = f"SELECT COUNT(*) as cnt FROM ({union_select})"
            query_params = params + params + [start_row, end_row]
            count_params = params + params
            cache_key = ("all", tuple(params), page, page_size)

        try:
            rows = await self._fb.cached_execute(cache_key, final_query, query_params)
            count_rows = await self._fb.cached_execute(
                ("count", *cache_key[1:]), final_count_query, count_params
            )
        except Exception as exc:
            log(f"❌ Firebird query failed: {exc}")
            return None

        # Normalize employee descriptions using the 4 digit code map
        for row in rows:
            row["description"] = map_employee_by_code(row.get("description"))

        total_count = int(count_rows[0]["cnt"]) if count_rows else 0
        total_pages = int((total_count + page_size - 1) / page_size)

        log(f"🔎 Firebird {total_count} records (page {page}/{total_pages})")

        return {
            "items": rows,
            "total": sum(int(r.get("kredit") or 0) for r in rows),
            "count": total_count,
            "avg": (
                float(sum(int(r.get("kredit") or 0) for r in rows) / len(rows))
                if rows
                else 0
            ),
            "page": page,
            "pages": total_pages,
        }

    async def get_sales_details(
        self,
        date_from: str | None = None,
        date_to: str | None = None,
        creater_ids: list[str] | None = None,
        user_ids: list[str] | None = None,
        folder_ids: list[str] | None = None,
        code_substr: str | None = None,
        name_substr: str | None = None,
        min_kredit: float | None = None,
        max_kredit: float | None = None,
        doc_num_substr: str | None = None,
        item_type: str | None = None,
        page: int = 1,
        page_size: int = 50,
    ) -> dict:
        if self._fb:
            return await self._firebird_details(
                date_from,
                date_to,
                creater_ids or [],
                user_ids or [],
                folder_ids or [],
                code_substr,
                name_substr,
                min_kredit,
                max_kredit,
                doc_num_substr,
                item_type,
                page,
                page_size,
            ) or {
                "items": [],
                "total": 0,
                "count": 0,
                "avg": 0,
                "page": page,
                "pages": 0,
            }

        df = await self._ensure_details_df()
        if df is None:
            return {
                "items": [],
                "total": 0,
                "count": 0,
                "avg": 0,
                "page": page,
                "pages": 0,
            }

        filtered = df
        if date_from:
            dt_from = pd.to_datetime(date_from, errors="coerce", dayfirst=True)
            if not pd.isna(dt_from):
                filtered = filtered[filtered["period"] >= dt_from]

        if date_to:
            dt_to = pd.to_datetime(date_to, errors="coerce", dayfirst=True)
            if not pd.isna(dt_to):
                filtered = filtered[filtered["period"] <= dt_to]

        if name_substr:
            filtered = filtered[
                filtered["item"].astype(str).str.contains(name_substr, case=False, na=False)
            ]

        if code_substr:
            filtered = filtered[
                filtered["employee"].astype(str).str.contains(code_substr, case=False, na=False)
            ]

        if doc_num_substr:
            filtered = filtered[
                filtered["order_number"].astype(str).str.contains(doc_num_substr, case=False, na=False)
            ]

        if min_kredit is not None:
            filtered = filtered[filtered["cost"] >= float(min_kredit)]

        if max_kredit is not None:
            filtered = filtered[filtered["cost"] <= float(max_kredit)]

        page_size = max(1, min(500, int(page_size)))
        page = max(1, int(page))
        total_count = int(len(filtered))
        total_pages = int((total_count + page_size - 1) / page_size)
        start = (page - 1) * page_size
        end = start + page_size
        page_df = filtered.iloc[start:end].copy()

        # format period back to string
        page_df["period"] = page_df["period"].dt.strftime("%Y-%m-%d")

        total_cost = int(filtered["cost"].sum())
        avg = float(total_cost / total_count) if total_count else 0.0

        log(f"🔎 Filtered {total_count} records (page {page}/{total_pages})")

        return {
            "items": page_df.to_dict(orient="records"),
            "total": total_cost,
            "count": total_count,
            "avg": avg,
            "page": page,
            "pages": total_pages,
        }

    async def _firebird_rating(
        self,
        date_from: str | None,
        date_to: str | None,
        creater_ids: list[str] | None,
        user_ids: list[str] | None,
        folder_ids: list[str],
        item_type: str | None,
    ) -> list[dict] | None:
        """Return aggregated sales by employee from Firebird."""
        if not self._fb:
            return None

        filters = ["docs_order_history.status_id = 5"]
        params: list[Any] = []
        if date_from:
            filters.append("docs.doc_date >= ?")
            params.append(date_from)
        if date_to:
            filters.append("docs.doc_date <= ?")
            params.append(date_to)
        if creater_ids:
            placeholders = ",".join(["?"] * len(creater_ids))
            filters.append(f"docs_order.creater_id IN ({placeholders})")
            params.extend(creater_ids)
        if user_ids:
            placeholders = ",".join(["?"] * len(user_ids))
            filters.append(f"users.user_id IN ({placeholders})")
            params.extend(user_ids)
        if folder_ids:
            placeholders = ",".join(["?"] * len(folder_ids))
            filters.append(f"tovars_tbl.folder_id IN ({placeholders})")
            params.extend(folder_ids)

        where_clause = " AND ".join(filters)
        if where_clause:
            where_clause = "WHERE " + where_clause

        lines_from = (
            "FROM doc_order_lines "
            "INNER JOIN docs_order ON doc_order_lines.doc_order_id = docs_order.id "
            "INNER JOIN docs_order_history ON docs_order.id = docs_order_history.doc_order_id "
            "INNER JOIN docs ON docs_order.doc_id = docs.doc_id "
            "INNER JOIN contragents ON docs.contragent_id = contragents.contr_id "
            "INNER JOIN tovars_tbl ON doc_order_lines.tovar_id = tovars_tbl.tovar_id "
            "INNER JOIN users ON docs_order.creater_id = users.user_id "
        )

        services_from = (
            "FROM doc_order_services "
            "INNER JOIN docs_order ON doc_order_services.doc_order_id = docs_order.id "
            "INNER JOIN docs_order_history ON docs_order.id = docs_order_history.doc_order_id "
            "INNER JOIN docs ON docs_order.doc_id = docs.doc_id "
            "INNER JOIN contragents ON docs.contragent_id = contragents.contr_id "
            "INNER JOIN tovars_tbl ON doc_order_services.tovar_id = tovars_tbl.tovar_id "
            "INNER JOIN users ON docs_order.creater_id = users.user_id "
        )

        lines_where = f"{where_clause} {'AND' if where_clause else 'WHERE'} doc_order_lines.kredit <> 0"
        services_where = f"{where_clause} {'AND' if where_clause else 'WHERE'} doc_order_services.kredit <> 0"

        lines_select = (
            "SELECT users.description AS description, "
            "SUM(doc_order_lines.kredit) AS total "
            f"{lines_from} {lines_where} GROUP BY users.description"
        )

        services_select = (
            "SELECT users.description AS description, "
            "SUM(doc_order_services.kredit) AS total "
            f"{services_from} {services_where} GROUP BY users.description"
        )

        if item_type == 'cosmetics':
            final_query = f"{lines_select} ORDER BY total DESC"
            query_params = params
            cache_key = ("cosmetics", tuple(params))
        elif item_type == 'services':
            final_query = f"{services_select} ORDER BY total DESC"
            query_params = params
            cache_key = ("services", tuple(params))
        else:
            union_select = f"{lines_select} UNION ALL {services_select}"
            final_query = (
                "SELECT description, SUM(total) AS total FROM (" + union_select + ") "
                "GROUP BY description ORDER BY total DESC"
            )
            query_params = params + params
            cache_key = ("all", tuple(params))

        try:
            rows = await self._fb.cached_execute(cache_key, final_query, query_params)
        except Exception as exc:
            log(f"❌ Firebird rating query failed: {exc}")
            return None

        for row in rows:
            row["description"] = map_employee_by_code(row.get("description"))
            row["total"] = int(row.get("total") or 0)

        return rows

    async def get_sales_rating(
        self,
        date_from: str | None = None,
        date_to: str | None = None,
        creater_ids: list[str] | None = None,
        user_ids: list[str] | None = None,
        folder_ids: list[str] | None = None,
        item_type: str | None = None,
    ) -> list[dict]:
        """Return rating of employees by sales amount."""
        if self._fb:
            return await self._firebird_rating(
                date_from,
                date_to,
                creater_ids or [],
                user_ids or [],
                folder_ids or [],
                item_type,
            ) or []

        df = await self._ensure_details_df()
        if df is None:
            return []

        filtered = df
        if date_from:
            dt_from = pd.to_datetime(date_from, errors="coerce", dayfirst=True)
            if not pd.isna(dt_from):
                filtered = filtered[filtered["period"] >= dt_from]

        if date_to:
            dt_to = pd.to_datetime(date_to, errors="coerce", dayfirst=True)
            if not pd.isna(dt_to):
                filtered = filtered[filtered["period"] <= dt_to]

        grouped = (
            filtered.groupby("employee")["cost"].sum().reset_index().rename(columns={"cost": "total"})
        )

        grouped["total"] = grouped["total"].astype(int)
        grouped = grouped.sort_values("total", ascending=False)

        return grouped.to_dict(orient="records")

    async def get_employee_detailed(
        self, date_from: str | None = None, date_to: str | None = None
    ) -> dict:
        """Return aggregated employee stats for the period."""

        shift_counts: dict[str, int] = {}

        from datetime import date, timedelta

        if date_from and date_to:
            try:
                start = date.fromisoformat(date_from)
                end = date.fromisoformat(date_to)
            except Exception:
                start = end = None
        else:
            start = end = None

        months: set[str] = set()
        if start and end:
            cur = start.replace(day=1)
            while cur <= end:
                months.add(MONTHS_RU[cur.month - 1])
                cur = (cur.replace(day=28) + timedelta(days=4)).replace(day=1)

        if months:
            shift_counts = _count_shifts_from_excel(list(months))

        df = await self._ensure_details_df()
        if df is None:
            cosmetics: dict[str, float] = {}
        else:
            filtered = df
            if date_from:
                dt_from = pd.to_datetime(date_from, errors="coerce", dayfirst=True)
                if not pd.isna(dt_from):
                    filtered = filtered[filtered["period"] >= dt_from]
            if date_to:
                dt_to = pd.to_datetime(date_to, errors="coerce", dayfirst=True)
                if not pd.isna(dt_to):
                    filtered = filtered[filtered["period"] <= dt_to]
            cosmetics = filtered.groupby("employee")["cost"].sum().to_dict()

        repair: dict[str, float] = {}
        shoes_sum: dict[str, float] = {}
        shoes_cnt: dict[str, int] = {}
        if self._fb:
            repair_query = (
                "SELECT users.description AS description, "
                "SUM(doc_order_services.kredit) AS total "
                "FROM docs_order "
                "INNER JOIN doc_order_services ON docs_order.id = doc_order_services.doc_order_id "
                "INNER JOIN tovars_tbl ON doc_order_services.tovar_id = tovars_tbl.tovar_id "
                "INNER JOIN docs ON docs_order.doc_id = docs.doc_id "
                "INNER JOIN users ON docs_order.creater_id = users.user_id "
                "WHERE docs.doc_date BETWEEN ? AND ? "
                "AND tovars_tbl.folder_id IN (215,216,217,221,326,327,328,329,330,416,417,418,419,108401,108402,110409,110410,110411,210266,210267,210268,210269,210270,210271,210272,210273,210274,210275,210276,210277,210278,210279,210280,210281,210282,210283,210284,210285,210286,210287,210288,210289,210290,210291,210292,210293,210294,210295,210296,210297,210298,210299,210300,210301,210302,210303,210304,210305,210306,210307,210308,210309,210310,210311,210312,210313,210314,210315,210316,210317,210318,210319,210320,210321,210322,210323,210324,210325,210326,210327,210328,210329,210330,210331,210332,210333,210334,210335,210336,210337,210338,210339,210340,210341,210342,210343,210344,210345,210346,210347,210348,210349,210350,210351,210352,210353,210355,210356,210357,210358,210359,210360,210361,210363,210364,210365,210366,210377,210378,210379,210380,210381,210382,210383,210384,210385,210386,210387,210388,210389,210390,210391,210392,210393,210394,210395,210396,210397,210399) "
                "GROUP BY users.description"
            )
            rows = await self._fb.execute(repair_query, [date_from, date_to])
            for row in rows:
                name = map_employee_by_code(row.get("description"))
                repair[name] = float(row.get("total") or 0)

            shoes_query = (
                "SELECT users.description AS description, "
                "SUM(doc_order_services.kredit) AS total, COUNT(*) AS cnt "
                "FROM doc_order_services "
                "INNER JOIN docs_order ON doc_order_services.doc_order_id = docs_order.id "
                "INNER JOIN docs ON docs_order.doc_id = docs.doc_id "
                "INNER JOIN tovars_tbl ON doc_order_services.tovar_id = tovars_tbl.tovar_id "
                "INNER JOIN users ON docs_order.creater_id = users.user_id "
                "WHERE docs_order.date_out_fact BETWEEN ? AND ? AND tovars_tbl.code = '1' "
                "GROUP BY users.description"
            )
            rows = await self._fb.execute(shoes_query, [date_from, date_to])
            for row in rows:
                name = map_employee_by_code(row.get("description"))
                shoes_sum[name] = float(row.get("total") or 0)
                shoes_cnt[name] = int(row.get("cnt") or 0)

        employees = set(shift_counts) | set(cosmetics) | set(repair) | set(shoes_sum)
        result = []
        for name in employees:
            shifts = shift_counts.get(name, 0)
            cos_total = cosmetics.get(name, 0.0)
            rep_total = repair.get(name, 0.0)
            sh_sum = shoes_sum.get(name, 0.0)
            sh_cnt = shoes_cnt.get(name, 0)
            result.append(
                {
                    "employee": name,
                    "shifts": shifts,
                    "cosmetics_total": cos_total,
                    "cosmetics_avg": (cos_total / shifts if shifts else 0.0),
                    "repair_total": rep_total,
                    "repair_avg": (rep_total / shifts if shifts else 0.0),
                    "shoes_sum": sh_sum,
                    "shoes_count": sh_cnt,
                    "revenue_total": cos_total + rep_total + sh_sum,
                }
            )

        return {"items": result}

