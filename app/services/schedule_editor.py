"""Правка листа графика через настоящий Excel (COM-автоматизация).

Почему COM, а не openpyxl. График и ФОТ лежат в ОДНОМ листе одного файла:
колонки C..AG — дни месяца, правее — зарплата. И 414 из 642 формул листа
считаются от ячеек графика:

    AS3 = COUNTIF(C3:AG3,"М")+COUNTIF(C3:AG3,"Ц")+...      смены
    BW3 = IF(AND(C3="Ц",COUNTIF($C$3:$C$21,"Ц")=1),1000,0) доплаты

openpyxl не умеет вычислять формулы — он хранит их текст и при сохранении
теряет кэш значений, посчитанный Excel. Проверено на копии боевого файла:
после save() через openpyxl AS3 из 7 превращается в None, BW3 из 7000 в None.
А весь код (payroll_service, боты, schedule_service) читает именно кэш —
то есть зарплата по всей системе показалась бы пустой. Excel пересчитывает
формулы сам при сохранении, поэтому правка идёт через него.

Координаты ячейки резолвятся через openpyxl (быстро и безопасно), а COM
используется только на саму запись — чтобы Excel был запущен как можно
меньше времени.
"""
from __future__ import annotations

import logging
import os
import shutil
import threading
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from ..config import EXCEL_FILE
from ..core.constants import MONTHS_RU
from ..data.salon_repository import get_salon_repository

log = logging.getLogger(__name__)

# Excel COM однопоточен и не переносит параллельных обращений к одному файлу,
# поэтому правки строго по очереди.
_write_lock = threading.Lock()

# Сколько ждать Excel, прежде чем считать его зависшим и убивать процесс.
EXCEL_TIMEOUT_S = 60

# Сколько резервных копий держать. Файл небольшой, а история правок ФОТ
# дороже места на диске.
BACKUP_KEEP = 50


class ScheduleEditError(Exception):
    """Ожидаемая ошибка правки — текст показывается пользователю как есть."""


def _valid_codes() -> set[str]:
    return {s.code for s in get_salon_repository().list_salons() if s.code}


def _backup_dir() -> Path:
    # Рядом с самим файлом, а НЕ внутри app/: deploy.ps1 зеркалит app/ через
    # robocopy /MIR и снёс бы бэкапы при первом же деплое.
    d = Path(EXCEL_FILE).parent / "schedule_backups"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _make_backup() -> Path:
    src = Path(EXCEL_FILE)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = _backup_dir() / f"{src.stem}_{stamp}{src.suffix}"
    shutil.copy2(src, dst)

    backups = sorted(_backup_dir().glob(f"{src.stem}_*{src.suffix}"))
    for old in backups[:-BACKUP_KEEP]:
        try:
            old.unlink()
        except OSError:
            pass
    return dst


def is_open_elsewhere() -> bool:
    """Занят ли файл прямо сейчас — то есть открыт ли он в Excel у человека.

    Проверяется реальная блокировка файла, а НЕ наличие служебного `~$Имя.xlsx`
    рядом. Excel создаёт такой файл на время работы, но при аварийном закрытии
    оставляет его навсегда: в проде нашёлся `~$ФОТ админы 2026.xlsx` от
    30.12.2025 при полностью выключенном Excel. Проверка по его наличию
    заблокировала бы правки насовсем.
    """
    import msvcrt

    try:
        with open(EXCEL_FILE, "r+b") as f:
            try:
                msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, 1)
                msvcrt.locking(f.fileno(), msvcrt.LK_UNLCK, 1)
            except OSError:
                return True
    except OSError:
        return True
    return False


def _sheet_title(wb, month: int) -> str | None:
    """Имя листа месяца — та же логика, что и при чтении графика."""
    month_name = MONTHS_RU[month - 1]
    for candidate in (month_name, month_name.upper()):
        if candidate in wb.sheetnames:
            return candidate
    for title in wb.sheetnames:
        if title.upper().startswith(month_name.upper()):
            return title
    return None


def _resolve_cell(year: int, month: int, employee: str, day: int) -> tuple[str, int, int]:
    """(имя листа, строка, колонка) для клетки графика.

    Раскладка читается из файла, а не угадывается: у листов разных месяцев
    номера дней стоят то в первой строке, то во второй.
    """
    if not os.path.exists(EXCEL_FILE):
        raise ScheduleEditError("Файл графика не найден на сервере.")

    try:
        num_days = (date(year + month // 12, month % 12 + 1, 1)
                    - date(year, month, 1)).days
    except ValueError:
        raise ScheduleEditError("Некорректный месяц.")
    if not 1 <= day <= num_days:
        raise ScheduleEditError(f"В этом месяце нет {day}-го числа.")

    wb = load_workbook(EXCEL_FILE, data_only=True, read_only=True)
    try:
        title = _sheet_title(wb, month)
        if title is None:
            raise ScheduleEditError(
                f"В файле нет листа за {MONTHS_RU[month - 1]} — правка невозможна."
            )
        sheet = wb[title]

        day_col = None
        for col in range(1, sheet.max_column + 1):
            for row in (1, 2):
                raw = str(sheet.cell(row=row, column=col).value or "").strip()
                if raw == str(day):
                    day_col = col
                    break
            if day_col:
                break
        if day_col is None:
            raise ScheduleEditError(f"В листе не найдена колонка {day}-го числа.")

        emp_row = None
        for row in range(3, sheet.max_row + 1):
            raw = sheet.cell(row=row, column=1).value
            if raw and str(raw).strip() == employee:
                emp_row = row
                break
        if emp_row is None:
            raise ScheduleEditError(f"Сотрудник «{employee}» не найден в листе.")

        return title, emp_row, day_col
    finally:
        wb.close()


def _write_via_excel(title: str, row: int, col: int, value: str) -> None:
    """Записать значение и дать Excel пересчитать формулы ФОТ."""
    import pythoncom
    import win32com.client

    result: dict = {}

    def work():
        pythoncom.CoInitialize()
        excel = None
        wb = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(os.path.abspath(EXCEL_FILE))
            wb.Worksheets(title).Cells(row, col).Value = value or None
            excel.CalculateFull()
            wb.Save()
            result["ok"] = True
        except Exception as e:  # noqa: BLE001 — прокидываем наверх как есть
            result["error"] = str(e)
        finally:
            try:
                if wb is not None:
                    wb.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            pythoncom.CoUninitialize()

    t = threading.Thread(target=work, daemon=True)
    t.start()
    t.join(EXCEL_TIMEOUT_S)

    if t.is_alive():
        # Зависший Excel держит файл заблокированным, поэтому его нужно
        # снять принудительно — иначе следующая правка тоже не пройдёт.
        os.system("taskkill /F /IM EXCEL.EXE >nul 2>&1")
        raise ScheduleEditError(
            "Excel не ответил за отведённое время, правка отменена. "
            "Файл восстановлен из резервной копии."
        )
    if "error" in result:
        raise ScheduleEditError(f"Excel не смог сохранить файл: {result['error']}")


def set_schedule_cell(year: int, month: int, employee: str, day: int, code: str) -> dict:
    """Проставить код салона сотруднику на день. Пустой код очищает клетку."""
    code = (code or "").strip()
    if code and code not in _valid_codes():
        raise ScheduleEditError(
            f"Неизвестный код «{code}». Допустимые: {', '.join(sorted(_valid_codes()))}"
        )

    if is_open_elsewhere():
        raise ScheduleEditError(
            "Файл сейчас открыт в Excel — закройте его, иначе правка потеряется "
            "при вашем сохранении."
        )

    title, row, col = _resolve_cell(year, month, employee, day)

    with _write_lock:
        # Повторная проверка под замком: файл могли открыть, пока мы ждали
        # своей очереди.
        if is_open_elsewhere():
            raise ScheduleEditError("Файл открыт в Excel — правка отменена.")

        backup = _make_backup()
        try:
            _write_via_excel(title, row, col, code)
        except Exception:
            # Возвращаем файл в исходное состояние: полупосчитанный ФОТ хуже,
            # чем отменённая правка.
            try:
                shutil.copy2(backup, EXCEL_FILE)
            except OSError as restore_err:
                log.error("Не удалось восстановить график из %s: %s", backup, restore_err)
            raise

    log.info("Schedule cell set: %s %s day=%s -> %r", title, employee, day, code)
    return {"sheet": title, "row": row, "col": col, "code": code, "backup": backup.name}
