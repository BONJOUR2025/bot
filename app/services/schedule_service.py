from __future__ import annotations

import calendar
import os
from datetime import date
from typing import Dict, List

from openpyxl import load_workbook

from ..config import EXCEL_FILE
from ..core.constants import MONTHS_RU
from ..data.salon_repository import get_salon_repository
from ..schemas.schedule import SchedulePointOut

_WEEKDAY_SHORT = ["пн", "вт", "ср", "чт", "пт", "сб", "вс"]


def _get_points() -> Dict[str, str]:
    """Map salon code -> salon name, as used by the Excel schedule."""
    return {
        salon.code: salon.name
        for salon in get_salon_repository().list_salons()
        if salon.code
    }


def _open_sheet(year: int, month: int):
    """Return (workbook, sheet) for the given month, or (None, None)."""
    if not os.path.exists(EXCEL_FILE):
        return None, None
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
    except Exception:
        return None, None
    month_name = MONTHS_RU[month - 1]
    for candidate in (month_name, month_name.upper()):
        if candidate in wb.sheetnames:
            return wb, wb[candidate]
    for title in wb.sheetnames:
        if title.upper().startswith(month_name.upper()):
            return wb, wb[title]
    return wb, None


class ScheduleService:
    """Load schedule from Excel by day."""

    def __init__(self) -> None:
        pass

    async def get_schedule_by_day(
            self, date_str: str) -> List[SchedulePointOut]:
        """Return list of points and assigned employees for given date."""

        try:
            day_date = date.fromisoformat(date_str)
        except Exception:
            return []

        points = _get_points()
        _, sheet = _open_sheet(day_date.year, day_date.month)

        if sheet is None:
            return [
                SchedulePointOut(point=name, short=code, employee="")
                for code, name in points.items()
            ]

        day_col = None
        target = str(day_date.day)
        for col in range(1, sheet.max_column + 1):
            v1 = str(sheet.cell(row=1, column=col).value or "").strip()
            v2 = str(sheet.cell(row=2, column=col).value or "").strip()
            if v1 == target or v2 == target:
                day_col = col
                break

        if day_col is None:
            return [
                SchedulePointOut(point=name, short=code, employee="")
                for code, name in points.items()
            ]

        assignments: Dict[str, str] = {}
        for row in range(3, sheet.max_row + 1):
            code = str(sheet.cell(row=row, column=day_col).value or "").strip()
            if code not in points or code in assignments:
                continue
            employee_cell = sheet.cell(row=row, column=1).value
            employee_name = str(employee_cell).strip() if employee_cell else ""
            assignments[code] = employee_name
            if len(assignments) == len(points):
                break

        return [
            SchedulePointOut(
                point=name,
                short=code,
                employee=assignments.get(code, ""),
            )
            for code, name in points.items()
        ]

    async def get_schedule_month(self, year: int, month: int) -> dict:
        """Return full month schedule as {employees, days, points}.

        days: list of {day, date, weekday_short, is_weekend, assignments: {employee: code}}
        employees: ordered list of employee names found in the sheet
        points: {code: name} mapping
        """
        num_days = calendar.monthrange(year, month)[1]

        def _empty_days():
            return [
                {
                    "day": d,
                    "date": date(year, month, d).isoformat(),
                    "weekday_short": _WEEKDAY_SHORT[date(year, month, d).weekday()],
                    "is_weekend": date(year, month, d).weekday() >= 5,
                    "assignments": {},
                }
                for d in range(1, num_days + 1)
            ]

        points = _get_points()
        _, sheet = _open_sheet(year, month)
        if sheet is None:
            return {"employees": [], "days": _empty_days(), "points": points}

        # Map day number → column index (row 1 or row 2 holds day numbers)
        day_cols: Dict[int, int] = {}
        for col in range(1, sheet.max_column + 1):
            for row in (1, 2):
                raw = str(sheet.cell(row=row, column=col).value or "").strip()
                try:
                    d = int(raw)
                    if 1 <= d <= 31 and d not in day_cols:
                        day_cols[d] = col
                except ValueError:
                    pass

        # Read employees (rows 3+, column 1) and their daily codes
        employees: List[str] = []
        emp_schedules: Dict[str, Dict[int, str]] = {}

        for row in range(3, sheet.max_row + 1):
            raw = sheet.cell(row=row, column=1).value
            emp_name = str(raw).strip() if raw else ""
            if not emp_name or emp_name.lower() in ("имя", "name"):
                continue
            sched: Dict[int, str] = {}
            for d, col in day_cols.items():
                val = str(sheet.cell(row=row, column=col).value or "").strip()
                if val:
                    sched[d] = val
            if sched:
                employees.append(emp_name)
                emp_schedules[emp_name] = sched

        days = []
        for d in range(1, num_days + 1):
            dt = date(year, month, d)
            assignments = {
                emp: emp_schedules[emp][d]
                for emp in employees
                if d in emp_schedules.get(emp, {})
            }
            days.append({
                "day": d,
                "date": dt.isoformat(),
                "weekday_short": _WEEKDAY_SHORT[dt.weekday()],
                "is_weekend": dt.weekday() >= 5,
                "assignments": assignments,
            })

        return {"employees": employees, "days": days, "points": points}
