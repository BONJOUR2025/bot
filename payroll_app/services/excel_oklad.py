import re
from openpyxl import load_workbook

CODE_RE = re.compile(r"(\d{4})$")

def extract_code(name: str) -> str | None:
    name = (name or "").strip()
    m = CODE_RE.search(name)
    return m.group(1) if m else None

def read_oklads(excel_path: str, sheet_name: str) -> dict[str, float]:
    """
    Возвращает {emp_code: oklad}
    Берём:
      A3:A21 -> имя "Имя 1234"
      AU3:AU21 -> оклад
    """
    wb = load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Нет листа '{sheet_name}'. Есть: {wb.sheetnames}")

    ws = wb[sheet_name]
    result: dict[str, float] = {}

    for row in range(3, 22):  # 3..21
        name = ws[f"A{row}"].value
        oklad = ws[f"AU{row}"].value

        code = extract_code(str(name)) if name else None
        if not code:
            continue

        try:
            result[code] = float(oklad or 0)
        except Exception:
            result[code] = 0.0

    return result
