"""Regression test: a blank ОСН/ДОП cell for one employee must not 500 the
whole month's payroll for everyone else.

Found in production: Анастасия 2602 was added to the АВГУСТ sheet without her
rate cells filled in (a new admin, rates not entered yet). pandas.read_excel
reads a blank cell as float('nan'), and `nan or 0` evaluates to nan — NaN is
truthy in Python — so the old `float(r.get("ОСН", 0) or 0)` left main_rate as
nan. That propagated into base_salary/total_gross/total_net, and FastAPI's
JSON encoder raises ValueError on out-of-range floats, turning one employee's
missing data point into a 500 for the entire /payroll/calculate response.
"""
from __future__ import annotations

import math

from openpyxl import Workbook

from app.services.payroll_service import PayrollService


def _build_workbook(path, *, blank_rates_for: str | None = None):
    """Minimal АВГУСТ sheet matching payroll_service's expected layout:
    name in column A, oklad in AU, bonus in BW, ОСН/ДОП header on row 2 at
    columns 34/36 (matches the real workbook's layout)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "АВГУСТ"

    # The real workbook has "ИМЯ" on BOTH header rows (row 1 is a coarser
    # grouping row, row 2 the actual per-column header) — pandas reads with
    # header=1 (the second row), so "ИМЯ" must be there too, not just A1.
    ws["A1"] = "ИМЯ"
    ws["A2"] = "ИМЯ"
    from openpyxl.utils import get_column_letter
    osn_col = get_column_letter(34)
    osn_shifts_col = get_column_letter(35)
    dop_col = get_column_letter(36)
    dop_shifts_col = get_column_letter(37)
    ws[f"{osn_col}1"] = "Ставка"
    ws[f"{osn_col}2"] = "ОСН"
    ws[f"{osn_shifts_col}2"] = "ОСН."
    ws[f"{dop_col}2"] = "ДОП"
    ws[f"{dop_shifts_col}2"] = "ДОП."

    employees = [("Вера 0102", 3500, 4000), ("Анастасия 2602", 3350, 3850)]
    for i, (name, osn, dop) in enumerate(employees, start=3):
        ws[f"A{i}"] = name
        ws[f"AU{i}"] = 0  # oklad
        ws[f"BW{i}"] = 0  # excel_bonus
        if name == blank_rates_for:
            continue  # leave ОСН/ДОП cells genuinely empty, like production
        ws[f"{osn_col}{i}"] = osn
        ws[f"{dop_col}{i}"] = dop

    wb.save(path)


def test_blank_rate_cell_becomes_zero_not_nan(tmp_path, monkeypatch):
    path = tmp_path / "payroll.xlsx"
    _build_workbook(path, blank_rates_for="Анастасия 2602")
    # ОСН/ДОП are read through a *separate* `from app.config import EXCEL_FILE`
    # inside the function body, independent of self.excel_path — patch that
    # module attribute, not the constructor arg, or the pandas read silently
    # falls back to the real production data.xlsx (or errors, as it did before
    # this was patched correctly).
    monkeypatch.setattr("app.config.EXCEL_FILE", str(path))
    svc = PayrollService(excel_path=str(path))

    employees = svc._get_employees_from_excel("АВГУСТ")
    anastasia = next(e for e in employees if "Анастасия" in e["name"])

    assert anastasia["main_rate"] == 0.0
    assert anastasia["extra_rate"] == 0.0
    assert not math.isnan(anastasia["main_rate"])
    assert not math.isnan(anastasia["extra_rate"])


def test_other_employees_rates_are_unaffected_by_a_blank_neighbor(tmp_path, monkeypatch):
    path = tmp_path / "payroll.xlsx"
    _build_workbook(path, blank_rates_for="Анастасия 2602")
    monkeypatch.setattr("app.config.EXCEL_FILE", str(path))
    svc = PayrollService(excel_path=str(path))

    employees = svc._get_employees_from_excel("АВГУСТ")
    vera = next(e for e in employees if "Вера" in e["name"])

    assert vera["main_rate"] == 3500.0
    assert vera["extra_rate"] == 4000.0


def test_no_blank_cells_leaves_rates_untouched(tmp_path, monkeypatch):
    path = tmp_path / "payroll.xlsx"
    _build_workbook(path, blank_rates_for=None)
    monkeypatch.setattr("app.config.EXCEL_FILE", str(path))
    svc = PayrollService(excel_path=str(path))

    employees = svc._get_employees_from_excel("АВГУСТ")
    for e in employees:
        assert e["main_rate"] > 0
        assert e["extra_rate"] > 0
