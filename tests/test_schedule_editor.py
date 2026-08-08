"""Тесты правки графика.

Сам Excel здесь не запускается — проверяется всё, что происходит ДО него:
валидация, поиск координат клетки, защита от занятого файла, резервные копии
и откат при сбое. Именно эта часть решает, попадёт ли в файл ФОТ мусор.

Почему запись вообще идёт через Excel, а не openpyxl — см. модуль сервиса:
414 формул зарплаты считаются от ячеек графика, а openpyxl при сохранении
теряет их посчитанные значения.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.services import schedule_editor as se


@pytest.fixture
def book(tmp_path, monkeypatch):
    """Мини-копия боевого листа: дни в строке 1, сотрудники с третьей."""
    path = tmp_path / "ФОТ тест.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "АВГУСТ"
    ws.cell(row=1, column=1, value="ИМЯ")
    for day in range(1, 32):
        ws.cell(row=1, column=2 + day, value=day)
    ws.cell(row=3, column=1, value="Вера 0102")
    ws.cell(row=4, column=1, value="Арина 7272")
    ws.cell(row=3, column=3, value="Ц")
    wb.save(path)

    monkeypatch.setattr(se, "EXCEL_FILE", str(path))
    monkeypatch.setattr(se, "_valid_codes", lambda: {"Ц", "М", "Гп"})
    return path


@pytest.fixture
def no_excel(monkeypatch):
    """Подменяет саму запись — Excel в тестах не поднимаем."""
    calls = []
    monkeypatch.setattr(se, "_write_via_excel",
                        lambda title, row, col, value: calls.append((title, row, col, value)))
    return calls


class TestValidation:
    def test_unknown_code_rejected_before_touching_the_file(self, book, no_excel):
        with pytest.raises(se.ScheduleEditError, match="Неизвестный код"):
            se.set_schedule_cell(2026, 8, "Вера 0102", 1, "ZZZ")
        assert no_excel == []

    def test_unknown_employee_rejected(self, book, no_excel):
        with pytest.raises(se.ScheduleEditError, match="не найден"):
            se.set_schedule_cell(2026, 8, "Нет Такого", 1, "Ц")
        assert no_excel == []

    def test_day_outside_month_rejected(self, book, no_excel):
        with pytest.raises(se.ScheduleEditError, match="нет 99"):
            se.set_schedule_cell(2026, 8, "Вера 0102", 99, "Ц")
        assert no_excel == []

    def test_february_30th_rejected(self, book, no_excel):
        """Границу месяца считаем по календарю, а не по «31 всегда есть»."""
        with pytest.raises(se.ScheduleEditError, match="нет 30"):
            se.set_schedule_cell(2026, 2, "Вера 0102", 30, "Ц")
        assert no_excel == []

    def test_missing_sheet_rejected(self, book, no_excel):
        with pytest.raises(se.ScheduleEditError, match="нет листа"):
            se.set_schedule_cell(2026, 12, "Вера 0102", 1, "Ц")
        assert no_excel == []

    def test_empty_code_is_allowed_and_clears_the_cell(self, book, no_excel):
        se.set_schedule_cell(2026, 8, "Вера 0102", 1, "")
        assert no_excel == [("АВГУСТ", 3, 3, "")]


class TestCellResolution:
    def test_resolves_row_and_column(self, book, no_excel):
        se.set_schedule_cell(2026, 8, "Арина 7272", 5, "М")
        title, row, col = no_excel[0][:3]
        assert (title, row, col) == ("АВГУСТ", 4, 7)  # 5-е число → колонка 2+5

    def test_writes_only_into_day_columns(self, book, no_excel):
        """Колонка обязана быть найдена по номеру дня, а не вычислена
        смещением: правее графика в том же листе лежат формулы ФОТ, и
        промах туда затёр бы расчёт зарплаты."""
        for day in (1, 15, 31):
            no_excel.clear()
            se.set_schedule_cell(2026, 8, "Вера 0102", day, "Ц")
            assert no_excel[0][2] == 2 + day


class TestFileBusy:
    def test_busy_file_is_not_written(self, book, no_excel, monkeypatch):
        monkeypatch.setattr(se, "is_open_elsewhere", lambda: True)
        with pytest.raises(se.ScheduleEditError, match="открыт в Excel"):
            se.set_schedule_cell(2026, 8, "Вера 0102", 1, "Ц")
        assert no_excel == []

    def test_free_file_reports_not_busy(self, book):
        assert se.is_open_elsewhere() is False

    def test_stale_lock_file_does_not_block(self, book):
        """В проде рядом с файлом лежал `~$…xlsx` от 30.12.2025 при
        выключенном Excel — проверка по его наличию заблокировала бы правки
        навсегда, поэтому смотрим на реальную блокировку файла."""
        (book.parent / f"~${book.name}").write_text("stale")
        assert se.is_open_elsewhere() is False


class TestBackups:
    def test_backup_created_before_write(self, book, no_excel):
        se.set_schedule_cell(2026, 8, "Вера 0102", 1, "Ц")
        backups = list((book.parent / "schedule_backups").glob("*.xlsx"))
        assert len(backups) == 1

    def test_backup_rotation_keeps_the_limit(self, book, monkeypatch):
        monkeypatch.setattr(se, "BACKUP_KEEP", 3)
        for i in range(5):
            # Штампы времени с секундной точностью — разводим руками, иначе
            # быстрые подряд идущие копии перезаписали бы друг друга.
            monkeypatch.setattr(se, "datetime", _FixedClock(i))
            se._make_backup()
        backups = list((book.parent / "schedule_backups").glob("*.xlsx"))
        assert len(backups) == 3

    def test_file_restored_when_excel_fails(self, book, monkeypatch):
        """Полупосчитанный ФОТ хуже отменённой правки, поэтому при сбое
        Excel файл возвращается из резервной копии."""
        original = book.read_bytes()

        def boom(title, row, col, value):
            book.write_bytes(b"corrupted")
            raise se.ScheduleEditError("Excel упал")

        monkeypatch.setattr(se, "_write_via_excel", boom)
        with pytest.raises(se.ScheduleEditError):
            se.set_schedule_cell(2026, 8, "Вера 0102", 1, "Ц")
        assert book.read_bytes() == original


class _FixedClock:
    """Фиксированные «часы» для проверки ротации копий."""

    def __init__(self, offset: int):
        self._offset = offset

    def now(self):
        from datetime import datetime as _dt, timedelta
        return _dt(2026, 1, 1) + timedelta(seconds=self._offset)
